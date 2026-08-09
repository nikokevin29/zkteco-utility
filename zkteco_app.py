#!/usr/bin/env python3
"""
ZKTeco eFace10 Utility — CV RAJ
Split panel: kiri workflow, kanan viewer + history
Semua data disimpan di SQLite, tidak ada file temp eksternal
"""

import csv, os, threading, calendar, sqlite3, json, sys, time, ssl, urllib.request, urllib.error
from collections import defaultdict
from datetime import datetime, date, timedelta

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QDialog, QLabel, QPushButton, QLineEdit,
    QComboBox, QCheckBox, QRadioButton, QVBoxLayout, QHBoxLayout, QGridLayout,
    QFormLayout, QGroupBox, QTabWidget, QSplitter, QTableWidget, QTableWidgetItem,
    QHeaderView, QPlainTextEdit, QFrame, QMessageBox, QFileDialog, QProgressBar,
    QSystemTrayIcon, QMenu, QAbstractItemView, QSizePolicy, QGraphicsOpacityEffect)
from PySide6.QtCore import Qt, QObject, Signal, QTimer, QPropertyAnimation, QEasingCurve
from PySide6.QtGui import QIcon, QAction, QColor, QFont
from PySide6.QtNetwork import QLocalServer, QLocalSocket

APP_VERSION = "5.0.3"


def _install_ssl_trust_fix():
    """
    Avoid Windows trust-store path bugs (e.g. expired DST Root CA X3 still trusted)
    that break Let's Encrypt chains with: certificate verify failed: certificate has expired.

    Prefer Mozilla CA bundle via certifi so urllib HTTPS does not load the broken
    Windows root path. Safe no-op if certifi is unavailable.
    """
    try:
        import certifi
        cafile = certifi.where()
        if not cafile or not os.path.isfile(cafile):
            return
    except Exception:
        return

    _orig = ssl.create_default_context

    def _create_default_context(*args, **kwargs):
        # Only inject when caller did not provide an explicit CA source.
        if not kwargs.get('cafile') and not kwargs.get('capath') and not kwargs.get('cadata'):
            kwargs = dict(kwargs)
            kwargs['cafile'] = cafile
        return _orig(*args, **kwargs)

    ssl.create_default_context = _create_default_context
    ssl._create_default_https_context = _create_default_context


_install_ssl_trust_fix()
_INSTANCE_KEY = "ZKTecoUtilityCVRAJ_single"
# Frozen exe: data lives next to the exe, NOT next to __file__ (which points
# into the throwaway _MEIxxxx extraction dir on onefile builds).
_BASE = (os.path.dirname(sys.executable) if getattr(sys, 'frozen', False)
         else os.path.dirname(os.path.abspath(__file__)))
CONFIG_FILE = os.path.join(_BASE, "config.json")
DB_FILE     = os.path.join(_BASE, "absensi.db")

BULAN_ID = ["Januari","Februari","Maret","April","Mei","Juni",
            "Juli","Agustus","September","Oktober","November","Desember"]
HARI_ID  = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"]

DEFAULT_CONFIG = {
    "ip": "10.10.11.55", "port": "8088",
    "lang": "en", "theme": "light",
    "jam_masuk": "08:00", "jam_keluar": "16:00",
    "toleransi": 15, "auto_backup": False,
    "autostart": False, "live_autostart": False,
    # Silent mode: tray only, Windows login, auto pull+sync
    "silent_mode": False,
    "auto_pull_on_start": True,      # pull (+cloud sync) after PC login
    "auto_pull_interval_min": 60,    # 0=hanya saat start; >0=ulang tiap N menit
    "anomaly_recover": True, "anomaly_anchor": "",
    # first_last = earliest/latest per day (default face verify)
    # device_punch = hormati status punch mesin (0=in, 1=out, …)
    "punch_mode": "first_last",
    "comm_password": 0,
    # VST-laravel cloud sync (service.rejekiamerta.com)
    "cloud_sync_enabled": False,
    "cloud_api_url": "https://service.rejekiamerta.com/api/attendance",
    "cloud_api_token": "",
    "cloud_sync_after_pull": True,
    "user_map": {
        "1":"NICHOLAS","2":"SERLI","3":"TIA","4":"MISRO",
        "5":"LISA","6":"TUR","7":"SLAMET","8":"ARI",
        "9":"REFA","10":"SUKUR","11":"PUGUH"
    }
}

# ZK punch codes (common on attendance devices)
PUNCH_IN_CODES = {0, 4}   # check-in, OT-in
PUNCH_OUT_CODES = {1, 5}  # check-out, OT-out
LEAVE_TYPES = ('izin', 'cuti', 'dinas', 'sakit')

# ─────────────────────────────────────────────────────────────────────────────
# ANOMALY (clock-reset) DETECTION & RECOVERY
# The eFace10 has NO RTC battery (UPS-only power backup). On a power loss the
# device clock resets to year 2000, so punches made during the outage get
# stamped with bogus year-2000 dates instead of the real date. These helpers
# detect those records and remap them back onto real calendar dates.
# ─────────────────────────────────────────────────────────────────────────────
ANOMALY_YEAR = 2000   # timestamp.year <= this  → clock-reset anomaly record

def is_anomaly_ts(ts):
    return ts is None or ts.year <= ANOMALY_YEAR

def _sec_of_day(ts):
    return ts.hour*3600 + ts.minute*60 + ts.second

def find_gaps(normal_recs, min_len=1):
    """List of (start, length_days) for calendar gaps in normal data, longest first."""
    dates = sorted({r['timestamp'].date() for r in normal_recs if r.get('timestamp')})
    if not dates:
        return []
    dset = set(dates)
    d = dates[0]
    gaps = []
    while d <= dates[-1]:
        if d not in dset and (d - timedelta(days=1)) in dset:
            s = d
            n = 0
            while d not in dset and d <= dates[-1] + timedelta(days=1):
                d += timedelta(days=1)
                n += 1
            if n >= min_len:
                gaps.append((s, n))
        else:
            d += timedelta(days=1)
    gaps.sort(key=lambda g: -g[1])
    return gaps


def find_gap_start(normal_recs, fallback=None, min_significant=2):
    """Start of the longest gap. Softened: if several significant gaps (or none),
    return None so caller can require a manual anchor instead of guessing."""
    gaps = find_gaps(normal_recs, min_len=1)
    if not gaps:
        if fallback:
            return fallback
        # no normal calendar — cannot auto-guess
        return None
    significant = [g for g in gaps if g[1] >= min_significant]
    if len(significant) > 1:
        # ambiguous: multiple multi-day holes (outage vs holiday/cuti)
        return None
    if significant:
        return significant[0][0]
    # only 1-day holes — still use longest, but single short gap is OK
    return gaps[0][0]

def remap_anomalies(anomaly_recs, anchor_date, jam_masuk='08:00', jam_keluar='16:00'):
    """Remap year-2000 records onto consecutive real dates starting at anchor_date
    (one fake day → one real day). Times within each day are linearly rescaled into
    a plausible work window [check-in − 30 min .. check-out + 60 min] so recovered
    rows look like normal days. Who-was-present and the punch order are REAL; the
    exact minute is approximate because the original clock was corrupted.
    Returns new dicts with corrected 'timestamp' plus 'recovered'=True and 'orig_ts'."""
    try:
        hi, mi = (int(x) for x in jam_masuk.split(':'))
        ho, mo = (int(x) for x in jam_keluar.split(':'))
    except Exception:
        hi, mi, ho, mo = 8, 0, 16, 0
    win_s = hi*3600 + mi*60 - 30*60
    win_e = ho*3600 + mo*60 + 60*60
    from collections import defaultdict
    byfake = defaultdict(list)
    for r in anomaly_recs:
        if r.get('timestamp'):
            byfake[r['timestamp'].date()].append(r)
    out = []; used = set()
    for i, fday in enumerate(sorted(byfake)):
        real = anchor_date + timedelta(days=i)
        recs = byfake[fday]
        secs = [_sec_of_day(r['timestamp']) for r in recs]
        t0, t1 = min(secs), max(secs); span = (t1 - t0) or 1
        base = datetime(real.year, real.month, real.day)
        for r in sorted(recs, key=lambda x: x['timestamp']):
            frac = (_sec_of_day(r['timestamp']) - t0) / span
            nts = base + timedelta(seconds=int(win_s + frac*(win_e - win_s)))
            while nts in used:               # keep timestamps unique (DB constraint)
                nts += timedelta(seconds=1)
            used.add(nts)
            nr = dict(r); nr['timestamp'] = nts
            nr['recovered'] = True; nr['orig_ts'] = r['timestamp']
            out.append(nr)
    return out

# ── Cross-platform opener ─────────────────────────────────────────────────────
import subprocess as _sp
def _open_path(path):
    try:
        if sys.platform == 'win32': os.startfile(path)
        elif sys.platform == 'darwin': _sp.Popen(['open', path])
        else: _sp.Popen(['xdg-open', path])
    except Exception: pass

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, encoding='utf-8') as f: cfg = json.load(f)
            for k,v in DEFAULT_CONFIG.items():
                if k not in cfg: cfg[k] = v
            return cfg
        except: pass
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

# ─────────────────────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        uid INTEGER, nama TEXT,
        timestamp TEXT,
        punch INTEGER, pulled_at TEXT,
        recovered INTEGER DEFAULT 0,
        UNIQUE(uid, timestamp)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        uid INTEGER PRIMARY KEY, nama TEXT,
        card_id TEXT, updated_at TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS pull_sessions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pulled_at TEXT,
        record_count INTEGER,
        new_count INTEGER,
        device_ip TEXT,
        note TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS excel_snapshots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id INTEGER,
        created_at TEXT,
        label TEXT,
        filter_year INTEGER,
        filter_month INTEGER,
        data BLOB
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS leave_days (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        uid INTEGER NOT NULL,
        tanggal TEXT NOT NULL,
        jenis TEXT NOT NULL,
        note TEXT DEFAULT '',
        UNIQUE(uid, tanggal)
    )''')
    # Company-wide holidays (manual only — weekend is NOT auto-libur)
    c.execute('''CREATE TABLE IF NOT EXISTS holidays (
        tanggal TEXT PRIMARY KEY,
        note TEXT DEFAULT '',
        created_at TEXT
    )''')
    _migrate_attendance_schema(conn)
    conn.commit()
    conn.close()


def _migrate_attendance_schema(conn):
    """Migrate legacy UNIQUE(timestamp) → UNIQUE(uid,timestamp); add recovered col."""
    c = conn.cursor()
    cols = {r[1] for r in c.execute('PRAGMA table_info(attendance)')}
    if not cols:
        return
    if 'recovered' not in cols:
        try:
            c.execute('ALTER TABLE attendance ADD COLUMN recovered INTEGER DEFAULT 0')
        except sqlite3.OperationalError:
            pass
    # Detect old global unique on timestamp only
    idx = list(c.execute("PRAGMA index_list(attendance)"))
    need_rebuild = False
    for _, name, unique, *_ in idx:
        if not unique:
            continue
        info = list(c.execute(f'PRAGMA index_info("{name}")'))
        col_names = [r[2] for r in info]
        if col_names == ['timestamp']:
            need_rebuild = True
            break
    # Also rebuild if no unique on (uid, timestamp)
    has_pair = False
    for _, name, unique, *_ in idx:
        if not unique:
            continue
        info = list(c.execute(f'PRAGMA index_info("{name}")'))
        col_names = [r[2] for r in info]
        if col_names == ['uid', 'timestamp'] or set(col_names) == {'uid', 'timestamp'}:
            has_pair = True
    if not has_pair:
        need_rebuild = True
    if not need_rebuild:
        return
    c.execute('''CREATE TABLE IF NOT EXISTS attendance_new (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        uid INTEGER, nama TEXT,
        timestamp TEXT,
        punch INTEGER, pulled_at TEXT,
        recovered INTEGER DEFAULT 0,
        UNIQUE(uid, timestamp)
    )''')
    c.execute('''INSERT OR IGNORE INTO attendance_new
                 (uid, nama, timestamp, punch, pulled_at, recovered)
                 SELECT uid, nama, timestamp, punch, pulled_at,
                        COALESCE(recovered, 0) FROM attendance''')
    c.execute('DROP TABLE attendance')
    c.execute('ALTER TABLE attendance_new RENAME TO attendance')


def db_insert_attendance(rows):
    """Insert rows. Returns (inserted, skipped_dup)."""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ins = 0
    skip = 0
    for r in rows:
        ts = r['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if hasattr(r['timestamp'], 'strftime') else str(r['timestamp'])
        rec = 1 if r.get('recovered') else 0
        try:
            c.execute(
                'INSERT INTO attendance (uid,nama,timestamp,punch,pulled_at,recovered) VALUES (?,?,?,?,?,?)',
                (r['uid'], r['nama'], ts, r.get('punch', 0), now, rec),
            )
            ins += 1
        except sqlite3.IntegrityError:
            skip += 1
    conn.commit()
    conn.close()
    return ins, skip


def db_query_attendance(year=None, month=None):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    q = ("SELECT uid,nama,timestamp,punch,COALESCE(recovered,0) FROM attendance "
         "WHERE strftime('%Y',timestamp)>'2000'")
    args = []
    if year:
        q += " AND strftime('%Y',timestamp)=?"
        args.append(str(year))
    if month:
        q += " AND strftime('%m',timestamp)=?"
        args.append(f"{month:02d}")
    q += " ORDER BY timestamp"
    c.execute(q, args)
    rows = [{
        'uid': r[0], 'nama': r[1],
        'timestamp': datetime.strptime(r[2], '%Y-%m-%d %H:%M:%S'),
        'punch': r[3], 'recovered': bool(r[4]),
    } for r in c.fetchall()]
    conn.close()
    return rows

def db_count():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM attendance WHERE strftime('%Y',timestamp)>'2000'")
    n = c.fetchone()[0]; conn.close(); return n

def db_upsert_users(users):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for u in users:
        c.execute('INSERT OR REPLACE INTO users (uid,nama,card_id,updated_at) VALUES (?,?,?,?)',
                  (u['uid'], u['nama'], u.get('card_id',''), now))
    conn.commit(); conn.close()

def db_add_pull_session(record_count, new_count, device_ip, note=''):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute('INSERT INTO pull_sessions (pulled_at,record_count,new_count,device_ip,note) VALUES (?,?,?,?,?)',
              (now, record_count, new_count, device_ip, note))
    sid = c.lastrowid
    conn.commit(); conn.close()
    return sid

def db_get_pull_sessions():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT id,pulled_at,record_count,new_count,device_ip FROM pull_sessions ORDER BY id DESC')
    rows = c.fetchall(); conn.close()
    return rows

def db_delete_pull_session(sid):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM pull_sessions WHERE id=?', (sid,))
    c.execute('DELETE FROM excel_snapshots WHERE session_id=?', (sid,))
    conn.commit(); conn.close()

def db_save_excel_snapshot(session_id, label, year, month, data_bytes):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute('INSERT INTO excel_snapshots (session_id,created_at,label,filter_year,filter_month,data) VALUES (?,?,?,?,?,?)',
              (session_id, now, label, year, month or 0, data_bytes))
    sid = c.lastrowid
    conn.commit(); conn.close()
    return sid

def db_get_excel_snapshots():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''SELECT e.id, e.label, e.created_at, e.filter_year, e.filter_month,
                        p.pulled_at, p.record_count
                 FROM excel_snapshots e
                 LEFT JOIN pull_sessions p ON e.session_id=p.id
                 ORDER BY e.id DESC''')
    rows = c.fetchall(); conn.close()
    return rows

def db_load_excel_snapshot(snap_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('SELECT data,label FROM excel_snapshots WHERE id=?', (snap_id,))
    row = c.fetchone(); conn.close()
    return row  # (bytes, label)

def db_delete_excel_snapshot(snap_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('DELETE FROM excel_snapshots WHERE id=?', (snap_id,))
    conn.commit(); conn.close()


# ── Leave (izin / cuti / dinas / sakit) — lokal di absensi.db ────────────────
def db_set_leave(uid, tanggal, jenis, note=''):
    jenis = jenis.lower().strip()
    if jenis not in LEAVE_TYPES:
        raise ValueError(f'jenis must be one of {LEAVE_TYPES}')
    tgl = tanggal if isinstance(tanggal, str) else tanggal.strftime('%Y-%m-%d')
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        'INSERT OR REPLACE INTO leave_days (uid, tanggal, jenis, note) VALUES (?,?,?,?)',
        (int(uid), tgl, jenis, note or ''),
    )
    conn.commit()
    conn.close()


def db_delete_leave(uid, tanggal):
    tgl = tanggal if isinstance(tanggal, str) else tanggal.strftime('%Y-%m-%d')
    conn = sqlite3.connect(DB_FILE)
    conn.execute('DELETE FROM leave_days WHERE uid=? AND tanggal=?', (int(uid), tgl))
    conn.commit()
    conn.close()


def db_list_leaves(year=None, month=None):
    conn = sqlite3.connect(DB_FILE)
    try:
        q = 'SELECT uid, tanggal, jenis, note FROM leave_days WHERE 1=1'
        args = []
        if year:
            q += " AND strftime('%Y', tanggal)=?"
            args.append(str(year))
        if month:
            q += " AND strftime('%m', tanggal)=?"
            args.append(f"{month:02d}")
        q += ' ORDER BY tanggal, uid'
        rows = conn.execute(q, args).fetchall()
    except sqlite3.OperationalError:
        rows = []  # table not migrated yet
    conn.close()
    return [{'uid': r[0], 'tanggal': r[1], 'jenis': r[2], 'note': r[3]} for r in rows]


def db_leave_map(year=None, month=None):
    """{(uid, date): jenis}"""
    m = {}
    for r in db_list_leaves(year, month):
        d = datetime.strptime(r['tanggal'], '%Y-%m-%d').date()
        m[(r['uid'], d)] = r['jenis']
    return m


def db_set_holiday(tanggal, note=''):
    """Mark a calendar date as company holiday (all employees)."""
    tgl = tanggal if isinstance(tanggal, str) else tanggal.strftime('%Y-%m-%d')
    datetime.strptime(tgl, '%Y-%m-%d')  # validate
    conn = sqlite3.connect(DB_FILE)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute(
        'INSERT OR REPLACE INTO holidays (tanggal, note, created_at) VALUES (?,?,?)',
        (tgl, note or '', now),
    )
    conn.commit()
    conn.close()


def db_delete_holiday(tanggal):
    tgl = tanggal if isinstance(tanggal, str) else tanggal.strftime('%Y-%m-%d')
    conn = sqlite3.connect(DB_FILE)
    conn.execute('DELETE FROM holidays WHERE tanggal=?', (tgl,))
    conn.commit()
    conn.close()


def db_list_holidays(year=None, month=None):
    conn = sqlite3.connect(DB_FILE)
    try:
        q = 'SELECT tanggal, note FROM holidays WHERE 1=1'
        args = []
        if year:
            q += " AND strftime('%Y', tanggal)=?"
            args.append(str(year))
        if month:
            q += " AND strftime('%m', tanggal)=?"
            args.append(f'{month:02d}')
        q += ' ORDER BY tanggal'
        rows = conn.execute(q, args).fetchall()
    except sqlite3.OperationalError:
        rows = []
    conn.close()
    return [{'tanggal': r[0], 'note': r[1] or ''} for r in rows]


def db_holiday_set(year=None, month=None):
    """set of date objects marked as holiday."""
    s = set()
    for r in db_list_holidays(year, month):
        s.add(datetime.strptime(r['tanggal'], '%Y-%m-%d').date())
    return s


# ─────────────────────────────────────────────────────────────────────────────
# CLOUD SYNC → VST-laravel (POST /api/attendance/sync)
# ─────────────────────────────────────────────────────────────────────────────
def normalize_cloud_token(token):
    """Bersihkan token dari paste error (Bearer prefix, spasi, BOM, quotes)."""
    t = (token or '').strip().strip('"').strip("'")
    t = t.replace('\ufeff', '').replace('\u200b', '').replace('\u00a0', ' ')
    # hapus semua whitespace (paste multi-baris / spasi di tengah)
    t = ''.join(t.split())
    if t.lower().startswith('bearer'):
        t = t[6:].lstrip(':').lstrip()
    return t


def normalize_cloud_base_url(url):
    """Normalisasi base URL agar selalu .../api/attendance (tanpa /sync)."""
    u = (url or '').strip().rstrip('/')
    if u.lower().endswith('/sync'):
        u = u[:-5].rstrip('/')
    return u


def build_cloud_payload(cfg, punches=None, year=None, month=None):
    """Build JSON body for VST attendance sync API."""
    um = cfg.get('user_map') or {}
    employees = [
        {'uid': int(k), 'name': str(v)}
        for k, v in um.items() if str(k).isdigit() and v
    ]
    if punches is None:
        punches = db_query_attendance(year, month)
    punch_payload = []
    for r in punches:
        ts = r['timestamp']
        if hasattr(ts, 'strftime'):
            ts = ts.strftime('%Y-%m-%d %H:%M:%S')
        punch_payload.append({
            'uid': int(r['uid']),
            'timestamp': str(ts),
            'punch': int(r.get('punch') or 0),
            'nama': r.get('nama') or '',
            'recovered': bool(r.get('recovered')),
        })
    leaves = db_list_leaves(year, month)
    leave_payload = [
        {
            'uid': int(l['uid']),
            'tanggal': l['tanggal'],
            'jenis': l['jenis'],
            'note': l.get('note') or '',
        }
        for l in leaves
    ]
    holidays = db_list_holidays(year, month)
    holiday_payload = [
        {'tanggal': h['tanggal'], 'note': h.get('note') or ''}
        for h in holidays
    ]
    return {
        'employees': employees,
        'punches': punch_payload,
        'leaves': leave_payload,
        'holidays': holiday_payload,
    }


def cloud_credentials_ok(cfg, log=None):
    """
    Tes credential ke VST (GET /employees).
    Returns dict {ok, status, message}. Raises RuntimeError jika config kosong.
    """
    def _log(msg):
        if log:
            log(msg)

    base = normalize_cloud_base_url(cfg.get('cloud_api_url') or '')
    token = normalize_cloud_token(cfg.get('cloud_api_token') or '')
    if not base or not token:
        raise RuntimeError('Cloud API URL / token kosong. Isi di Settings → Cloud Sync.')

    url = base + '/employees'
    req = urllib.request.Request(
        url, method='GET',
        headers={
            'Accept': 'application/json',
            'Authorization': f'Bearer {token}',
            'X-Attendance-Token': token,  # fallback jika proxy strip Authorization
            'User-Agent': f'ZKTeco-Utility/{APP_VERSION}',
        },
    )
    _log(f'☁ Tes credential → {url} ...')
    try:
        with urllib.request.urlopen(req, timeout=30, context=ssl.create_default_context()) as resp:
            raw = resp.read().decode('utf-8', errors='replace')
            data = json.loads(raw) if raw else {}
            n = len(data.get('employees') or [])
            msg = f'Credential OK — {n} karyawan di VST'
            _log(f'✓ {msg}')
            return {'ok': True, 'status': resp.status, 'message': msg, 'employees': n}
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8', errors='replace')[:400]
        hint = _cloud_auth_hint(e.code, err_body)
        raise RuntimeError(f'HTTP {e.code}: {err_body or e.reason}{hint}') from e
    except urllib.error.URLError as e:
        raise RuntimeError(f'Network: {e.reason}') from e


def _cloud_auth_hint(code, body=''):
    b = (body or '').lower()
    if code == 401:
        return (
            '\n→ Token ditolak server. Di web VST: Pengaturan → Sync ZKTeco Desktop → '
            'salin token terbaru (jangan pakai prefix "Bearer "). '
            'Generate Ulang Token mematikan token lama di desktop.'
        )
    if code == 404:
        return (
            '\n→ URL salah. Base harus seperti: '
            'https://service.rejekiamerta.com/api/attendance (tanpa /sync di ujung).'
        )
    if code == 422 or 'validation' in b:
        return '\n→ Payload tidak valid (bukan masalah credential).'
    return ''


def cloud_sync(cfg, punches=None, year=None, month=None, log=None):
    """POST employees + punches + leaves to VST API. Returns result dict or raises."""
    def _log(msg):
        if log:
            log(msg)

    if not cfg.get('cloud_sync_enabled'):
        raise RuntimeError('Cloud sync disabled (Settings → Cloud Sync).')
    base = normalize_cloud_base_url(cfg.get('cloud_api_url') or '')
    token = normalize_cloud_token(cfg.get('cloud_api_token') or '')
    # persist cleaned credentials so next save/load stays clean
    cfg['cloud_api_url'] = base
    cfg['cloud_api_token'] = token
    if not base or not token:
        raise RuntimeError('Cloud API URL / token kosong. Isi di Settings → Cloud Sync.')

    url = base + '/sync'
    body = build_cloud_payload(cfg, punches=punches, year=year, month=month)
    data = json.dumps(body).encode('utf-8')
    req = urllib.request.Request(
        url, data=data, method='POST',
        headers={
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'Authorization': f'Bearer {token}',
            'X-Attendance-Token': token,  # fallback jika proxy strip Authorization
            'User-Agent': f'ZKTeco-Utility/{APP_VERSION}',
        },
    )
    _log(f'☁ Sync → {url}  ({len(body["punches"])} punch, '
         f'{len(body["employees"])} emp, {len(body["leaves"])} leave, '
         f'{len(body.get("holidays", []))} libur) ...')
    try:
        with urllib.request.urlopen(req, timeout=60, context=ssl.create_default_context()) as resp:
            raw = resp.read().decode('utf-8', errors='replace')
            result = json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8', errors='replace')[:400]
        hint = _cloud_auth_hint(e.code, err_body)
        raise RuntimeError(f'HTTP {e.code}: {err_body or e.reason}{hint}') from e
    except urllib.error.URLError as e:
        raise RuntimeError(f'Network: {e.reason}') from e

    _log(
        f'✓ Cloud sync OK — punches new={result.get("punches_new", "?")} '
        f'skip={result.get("punches_skipped", "?")} '
        f'emp={result.get("employees_upserted", "?")} '
        f'leave={result.get("leaves_upserted", "?")}'
    )
    return result


def resolve_in_out(taps, punch_mode='first_last'):
    """From list of (datetime, punch) → (masuk_dt|None, keluar_dt|None, tap_total)."""
    if not taps:
        return None, None, 0
    taps = sorted(taps, key=lambda x: x[0])
    n = len(taps)
    if punch_mode == 'device_punch':
        ins = [t for t, p in taps if int(p) in PUNCH_IN_CODES]
        outs = [t for t, p in taps if int(p) in PUNCH_OUT_CODES]
        if not ins and not outs:
            # device always sends 0 — fall back
            masuk = taps[0][0]
            keluar = taps[-1][0] if n > 1 else None
            return masuk, keluar, n
        masuk = ins[0] if ins else None
        keluar = outs[-1] if outs else None
        return masuk, keluar, n
    # first_last
    masuk = taps[0][0]
    keluar = taps[-1][0] if n > 1 else None
    return masuk, keluar, n


# Payroll week: Sunday → Saturday (payroll cut typically every Saturday)
WEEKDAY_SUN_FIRST = ['Minggu', 'Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu']


def sunday_of_week(d):
    """Calendar date of Sunday that starts the week containing d."""
    # Python: Mon=0 … Sun=6 → offset to previous/same Sunday
    return d - timedelta(days=(d.weekday() + 1) % 7)


def build_month_week_grid(year, month, day_status_map, today=None):
    """Build Sun–Sat week rows for one month.

    day_status_map: {date: {'status','masuk','keluar',...}} for ONE employee
    Returns list of week dicts:
      {'week': 1, 'label': 'M1 28/07–03/08', 'days': [cell|None × 7]}
    cell: {date, in_month, status, masuk, keluar, text}
    """
    if today is None:
        today = date.today()
    first = date(year, month, 1)
    _, days_in = calendar.monthrange(year, month)
    last = date(year, month, days_in)
    start = sunday_of_week(first)
    end = sunday_of_week(last) + timedelta(days=6)  # Saturday of last week
    weeks = []
    w = 0
    cur = start
    while cur <= end:
        w += 1
        days = []
        for i in range(7):
            d = cur + timedelta(days=i)
            in_month = d.month == month and d.year == year
            if not in_month:
                days.append({
                    'date': d, 'in_month': False, 'status': '',
                    'masuk': '', 'keluar': '', 'text': '',
                })
                continue
            info = day_status_map.get(d) or {}
            status = info.get('status', '')
            masuk = info.get('masuk', '-') or '-'
            keluar = info.get('keluar', '-') or '-'
            # future days in month: blank
            if d > today:
                text = ''
                status = ''
            elif status in ('HADIR', 'TELAT'):
                text = f'{masuk}\n{keluar}' if keluar != '-' else masuk
            elif status == 'NO_CHECKOUT':
                text = f'{masuk}\n—'
            elif status:
                text = status
            else:
                text = ''
            days.append({
                'date': d, 'in_month': True, 'status': status,
                'masuk': masuk, 'keluar': keluar, 'text': text,
            })
        label = (
            f'M{w} {days[0]["date"].strftime("%d/%m")}'
            f'–{days[6]["date"].strftime("%d/%m")}'
        )
        weeks.append({'week': w, 'label': label, 'days': days})
        cur += timedelta(days=7)
    return weeks


def build_employee_day_status_map(uid, year, month, rows, cfg, leave_map=None, holiday_set=None):
    """Full calendar day → status for one uid (fills ABSEN / LIBUR like payroll)."""
    if leave_map is None:
        leave_map = db_leave_map(year, month)
    if holiday_set is None:
        holiday_set = db_holiday_set(year, month)
    # filter rows for this uid
    uid_rows = [r for r in rows if int(r['uid']) == int(uid)]
    drows = compute_daily_rows(uid_rows, cfg, leave_map)
    by_day = {}
    for d in drows:
        if d['uid'] != int(uid):
            continue
        by_day[d['tanggal']] = d
    _, days_in = calendar.monthrange(year, month)
    today = date.today()
    user_map = {int(k): v for k, v in cfg.get('user_map', {}).items()}
    nama = user_map.get(int(uid), f'UID:{uid}')
    for day in range(1, days_in + 1):
        tgl = date(year, month, day)
        if tgl > today:
            continue
        if tgl in by_day:
            continue
        # leave-only already in compute_daily_rows; holidays + absen fill
        leave = leave_map.get((int(uid), tgl), '')
        if leave:
            st = leave.upper()
        elif tgl in holiday_set:
            st = 'LIBUR'
        else:
            st = 'ABSEN'
        by_day[tgl] = {
            'uid': int(uid), 'nama': nama, 'tanggal': tgl,
            'masuk': '-', 'keluar': '-', 'status': st, 'leave': leave,
        }
    return by_day


def compute_daily_rows(rows, cfg, leave_map=None):
    """Per-(uid,nama,tanggal) daily summary. Shared by Daily tab, Excel, payroll."""
    user_map = {int(k): v for k, v in cfg.get('user_map', {}).items()}
    std_in = datetime.strptime(cfg.get('jam_masuk', '08:00'), '%H:%M')
    std_out = datetime.strptime(cfg.get('jam_keluar', '16:00'), '%H:%M')
    tol_dt = std_in + timedelta(minutes=int(cfg.get('toleransi', 15)))
    punch_mode = cfg.get('punch_mode', 'first_last')
    if leave_map is None:
        leave_map = {}

    raw = {}  # (uid, nama, tgl) -> list[(ts, punch)]
    recovered_days = set()
    for r in rows:
        ts = r['timestamp']
        if isinstance(ts, str):
            ts = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')
        if ts.year <= ANOMALY_YEAR:
            continue
        uid = int(r['uid'])
        nama = user_map.get(uid, r.get('nama') or f"UID:{uid}")
        key = (uid, nama, ts.date())
        raw.setdefault(key, []).append((ts, int(r.get('punch') or 0)))
        if r.get('recovered'):
            recovered_days.add((uid, ts.date()))

    out = []
    for (uid, nama, tgl), taps in sorted(raw.items(), key=lambda kv: (kv[0][2], kv[0][1])):
        masuk, keluar, tap_total = resolve_in_out(taps, punch_mode)
        no_checkout = masuk is not None and keluar is None
        telat = 0
        if masuk:
            t = datetime.strptime(masuk.strftime('%H:%M'), '%H:%M')
            telat = int((t - std_in).total_seconds() // 60) if t > tol_dt else 0
        pulang_cepat = 0
        if keluar:
            t = datetime.strptime(keluar.strftime('%H:%M'), '%H:%M')
            if t < std_out:
                pulang_cepat = int((std_out - t).total_seconds() // 60)
        lembur = 0
        if keluar:
            t = datetime.strptime(keluar.strftime('%H:%M'), '%H:%M')
            lembur = int((t - std_out).total_seconds() // 60) if t > std_out else 0
        durasi = int((keluar - masuk).total_seconds() // 60) if (masuk and keluar) else 0
        leave = leave_map.get((uid, tgl), '')
        if leave:
            status = leave.upper()
        elif no_checkout:
            status = 'NO_CHECKOUT'
        elif telat > 0:
            status = 'TELAT'
        else:
            status = 'HADIR'
        out.append({
            'uid': uid, 'nama': nama, 'tanggal': tgl,
            'masuk': masuk.strftime('%H:%M') if masuk else '-',
            'keluar': keluar.strftime('%H:%M') if keluar else '-',
            'telat': telat, 'pulang_cepat': pulang_cepat, 'lembur': lembur,
            'durasi': durasi, 'tap_total': tap_total,
            'no_checkout': no_checkout,
            'recovered': (uid, tgl) in recovered_days,
            'leave': leave, 'status': status,
            'weekend': tgl.weekday() >= 5,
        })
    # days with leave but no punch
    seen = {(d['uid'], d['tanggal']) for d in out}
    for (uid, tgl), jenis in leave_map.items():
        if (uid, tgl) in seen:
            continue
        nama = user_map.get(uid, f'UID:{uid}')
        out.append({
            'uid': uid, 'nama': nama, 'tanggal': tgl,
            'masuk': '-', 'keluar': '-', 'telat': 0, 'pulang_cepat': 0, 'lembur': 0,
            'durasi': 0, 'tap_total': 0, 'no_checkout': False,
            'recovered': False, 'leave': jenis, 'status': jenis.upper(),
            'weekend': tgl.weekday() >= 5,
        })
    out.sort(key=lambda d: (d['tanggal'], d['nama']))
    return out


def generate_payroll_bytes(rows, cfg, year=None, month=None):
    """CSV payroll-friendly: one row per employee-day.

    Setiap hari kalender (sampai hari ini) untuk setiap UID di user_map.
    Sabtu/Minggu BUKAN libur otomatis — libur hanya dari daftar holidays
    (set manual). Tanpa tap & tanpa izin/cuti/sakit → ABSEN.
    """
    leave_map = db_leave_map(year, month)
    holiday_set = db_holiday_set(year, month)
    daily = compute_daily_rows(rows, cfg, leave_map)
    # Overlay: hari libur manual + fill ABSEN untuk hari kosong
    if year and month:
        user_map = {int(k): v for k, v in cfg.get('user_map', {}).items()}
        _, days_in = calendar.monthrange(year, month)
        # punch days that fall on holiday keep HADIR/etc unless leave
        for d in daily:
            if d['tanggal'] in holiday_set and not d.get('leave'):
                # tetap datang di hari libur → biarkan status hadir; jika ingin flag bisa LEBUR
                pass
        have = {(d['uid'], d['tanggal']) for d in daily}
        for uid, nama in user_map.items():
            for day in range(1, days_in + 1):
                tgl = date(year, month, day)
                if (uid, tgl) in have:
                    continue
                if tgl > date.today():
                    continue
                # LIBUR hanya jika tanggal di-set manual sebagai hari libur
                if tgl in holiday_set:
                    status = 'LIBUR'
                else:
                    status = 'ABSEN'
                daily.append({
                    'uid': uid, 'nama': nama, 'tanggal': tgl,
                    'masuk': '-', 'keluar': '-', 'telat': 0, 'pulang_cepat': 0,
                    'lembur': 0, 'durasi': 0, 'tap_total': 0,
                    'no_checkout': False, 'recovered': False,
                    'leave': '', 'status': status, 'weekend': tgl.weekday() >= 5,
                })
        daily.sort(key=lambda d: (d['tanggal'], d['nama']))

    import io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        'UID', 'Nama', 'Tanggal', 'Hari', 'Masuk', 'Keluar',
        'Telat_mnt', 'PulangCepat_mnt', 'Lembur_mnt', 'Durasi_mnt',
        'Taps', 'NoCheckout', 'Recovered', 'Status',
    ])
    hari = HARI_ID
    for d in daily:
        w.writerow([
            d['uid'], d['nama'], d['tanggal'].strftime('%Y-%m-%d'),
            hari[d['tanggal'].weekday()],
            d['masuk'], d['keluar'],
            d['telat'], d['pulang_cepat'], d['lembur'], d['durasi'],
            d['tap_total'],
            'Y' if d['no_checkout'] else '',
            'Y' if d['recovered'] else '',
            d['status'],
        ])
    return buf.getvalue().encode('utf-8-sig')

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATOR — returns bytes instead of saving to file
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel_bytes(rows, cfg):
    """Generate Excel and return as bytes (not saved to disk)."""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl tidak terinstall.")

    jam_masuk_std  = cfg.get('jam_masuk','08:00')
    jam_keluar_std = cfg.get('jam_keluar','16:00')
    toleransi      = int(cfg.get('toleransi', 15))
    user_map       = {int(k):v for k,v in cfg.get('user_map',{}).items()}
    lang           = cfg.get('lang','en')

    def F(h): return PatternFill('solid', fgColor=h)
    def Bs(st='thin', co='FFB0B0B0'):
        s=Side(style=st,color=co); return Border(left=s,right=s,top=s,bottom=s)
    def Bm(co='FF888888'):
        s=Side(style='medium',color=co); return Border(left=s,right=s,top=s,bottom=s)
    def fnt(sz=9, bold=False, color='FF000000', italic=False):
        return Font(name='Arial', size=sz, bold=bold, color=color, italic=italic)
    C  = Alignment(horizontal='center', vertical='center', wrap_text=False)
    CW = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L  = Alignment(horizontal='left',   vertical='center')
    R  = Alignment(horizontal='right',  vertical='center')

    CH='FF1A56DB'; CS='FF3B82F6'; CA='FF1E40AF'
    GBG='FFD1FAE5'; GTX='FF065F46'
    RBG='FFFEE2E2'; RTX='FF991B1B'
    YBG='FFFEF9C3'; YTX='FF854D0E'
    OBG='FFFED7AA'; OTX='FF9A3412'
    R1='FFFFFFFF'; R2='FFF0F4FF'; STR='FFE8F0FE'; BD='FFCBD5E1'; HT='FFFFFFFF'

    # month names
    MNAMES = (["","January","February","March","April","May","June",
               "July","August","September","October","November","December"]
              if lang=='en' else
              ["","Januari","Februari","Maret","April","Mei","Juni",
               "Juli","Agustus","September","Oktober","November","Desember"])
    DNAMES = (["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
              if lang=='en' else
              ["Sn","Sl","Rb","Km","Jm","Sb","Mg"])
    DFULL  = (["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
              if lang=='en' else
              ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"])

    # ── build data ────────────────────────────────────────────────────────────
    def _parse_ts(t):
        if isinstance(t, datetime): return t
        if isinstance(t, str):
            for fmt in ('%Y-%m-%d %H:%M:%S','%Y-%m-%d %H:%M'):
                try: return datetime.strptime(t, fmt)
                except: pass
        return None

    punch_mode = cfg.get('punch_mode', 'first_last')
    # leave map for period (optional — when generating from filtered rows)
    years = set()
    months = set()
    for r in rows:
        ts = _parse_ts(r.get('timestamp'))
        if ts and ts.year > 2000:
            years.add(ts.year); months.add(ts.month)
    leave_map = {}
    for y in years:
        for m in (months or [None]):
            leave_map.update(db_leave_map(y, m))

    _raw = {}  # (nama, tgl) -> list[(ts, punch, recovered)]
    for r in rows:
        ts = _parse_ts(r['timestamp'])
        if ts is None or ts.year <= 2000: continue
        nama = user_map.get(r['uid'], r.get('nama', f"UID:{r['uid']}"))
        tgl  = ts.date()
        _raw.setdefault((nama, tgl), []).append(
            (ts, int(r.get('punch') or 0), bool(r.get('recovered'))))

    if not _raw and not leave_map:
        raise RuntimeError("No data for selected period.")

    class _Row:
        __slots__ = ['nama','tanggal','masuk','keluar','tap','tap_total',
                     'jam_masuk','jam_keluar','terlambat','lembur',
                     'no_checkout','recovered','status','pulang_cepat']

    daily_list = []
    for (nama, tgl), taps in sorted(_raw.items()):
        pairs = [(t, p) for t, p, _ in taps]
        masuk, keluar, tap_total = resolve_in_out(pairs, punch_mode)
        row = _Row()
        row.nama = nama; row.tanggal = tgl
        row.masuk = masuk; row.keluar = keluar
        row.tap = 2 if (masuk and keluar) else (1 if masuk else 0)
        row.tap_total = tap_total
        row.jam_masuk = masuk.strftime('%H:%M') if masuk else '-'
        row.jam_keluar = keluar.strftime('%H:%M') if keluar else '-'
        row.terlambat = 0; row.lembur = 0; row.pulang_cepat = 0
        row.no_checkout = bool(masuk and not keluar)
        row.recovered = any(rec for _, _, rec in taps)
        row.status = 'NO_CHECKOUT' if row.no_checkout else 'HADIR'
        daily_list.append(row)

    class _DF:
        def __init__(self, lst): self._lst = lst
        def __iter__(self): return iter(self._lst)
        def __len__(self): return len(self._lst)
        def min_date(self): return min(r.tanggal for r in self._lst)
        def max_date(self): return max(r.tanggal for r in self._lst)
        def names(self): return sorted(set(r.nama for r in self._lst))
        def months(self): return sorted(set(r.tanggal.strftime('%Y-%m') for r in self._lst))
        def filter_month(self,m): return _DF([r for r in self._lst if r.tanggal.strftime('%Y-%m')==m])
        def filter_name(self,n):  return _DF([r for r in self._lst if r.nama==n])
        def filter_date(self,d):  return _DF([r for r in self._lst if r.tanggal==d])
        def filter_late(self):    return _DF([r for r in self._lst if r.terlambat>0])
        def sort_by(self,*keys):  return _DF(sorted(self._lst, key=lambda r: tuple(getattr(r,k) for k in keys)))

    daily = _DF(daily_list)

    std_in  = datetime.strptime(jam_masuk_std,'%H:%M')
    tol_dt  = std_in + timedelta(minutes=toleransi)
    std_out = datetime.strptime(jam_keluar_std,'%H:%M')

    for row in daily:
        try:
            if row.jam_masuk == '-':
                row.terlambat = 0
            else:
                t = datetime.strptime(row.jam_masuk, '%H:%M')
                row.terlambat = int((t - std_in).total_seconds() // 60) if t > tol_dt else 0
                if row.terlambat > 0 and not row.no_checkout:
                    row.status = 'TELAT'
        except Exception:
            row.terlambat = 0
        try:
            if row.jam_keluar == '-':
                row.lembur = 0
                row.pulang_cepat = 0
            else:
                t = datetime.strptime(row.jam_keluar, '%H:%M')
                row.lembur = int((t - std_out).total_seconds() // 60) if t > std_out else 0
                row.pulang_cepat = int((std_out - t).total_seconds() // 60) if t < std_out else 0
        except Exception:
            row.lembur = 0
            row.pulang_cepat = 0

    if len(daily) == 0:
        raise RuntimeError("No data for selected period.")
    months    = daily.months()
    all_names = daily.names()
    wb = Workbook()
    wb.remove(wb.active)

    # ── KARTU ABSENSI per BULAN ───────────────────────────────────────────────
    for month in months:
        yr2,mo2   = int(month[:4]),int(month[5:])
        _,days_in = calendar.monthrange(yr2,mo2)
        bln_label = f"{MNAMES[mo2]} {yr2}"
        mdata     = daily.filter_month(month)

        ws = wb.create_sheet(f'{MNAMES[mo2][:3]}{yr2}')
        ws.sheet_view.showGridLines=False
        ws.page_setup.orientation='landscape'; ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1

        TOT_COL=days_in+3; LAT_COL=days_in+4; OT_COL=days_in+5; AVG_COL=days_in+6
        LC=get_column_letter(days_in+6)

        ws.merge_cells(f'A1:{LC}1')
        ws['A1']=('CV REJEKI AMERTA JAYA — ATTENDANCE CARD' if lang=='en'
                  else 'CV REJEKI AMERTA JAYA — KARTU ABSENSI KARYAWAN')
        ws['A1'].font=fnt(13,True,HT); ws['A1'].fill=F(CH); ws['A1'].alignment=C
        ws.row_dimensions[1].height=22

        ws.merge_cells(f'A2:{LC}2')
        _std = ('Check-in' if lang=='en' else 'Masuk Std')
        _tol = ('Tolerance' if lang=='en' else 'Toleransi')
        ws['A2']=f"{bln_label.upper()}  |  {_std}: {jam_masuk_std}  |  {_tol}: {toleransi} min"
        ws['A2'].font=fnt(9,False,HT,italic=True); ws['A2'].fill=F(CS); ws['A2'].alignment=C
        ws.row_dimensions[2].height=15

        ws.merge_cells(f'A3:{LC}3')
        _printed = 'Printed:' if lang=='en' else 'Dicetak:'
        ws['A3']=f"{_printed} {datetime.now().strftime('%d %B %Y %H:%M')}"
        ws['A3'].font=fnt(8,False,'FF555555',italic=True); ws['A3'].fill=F('FFF8FAFF'); ws['A3'].alignment=R
        ws.row_dimensions[3].height=13

        ws.column_dimensions['A'].width=4; ws.column_dimensions['B'].width=15
        for d in range(1,days_in+1): ws.column_dimensions[get_column_letter(d+2)].width=5
        ws.column_dimensions[get_column_letter(TOT_COL)].width=5.5
        ws.column_dimensions[get_column_letter(LAT_COL)].width=9
        ws.column_dimensions[get_column_letter(OT_COL)].width=8
        ws.column_dimensions[get_column_letter(AVG_COL)].width=7

        HDR=4; ws.row_dimensions[HDR].height=28
        def hc(col,val,bg=CA):
            c=ws.cell(row=HDR,column=col,value=val)
            c.font=fnt(8,True,HT); c.fill=F(bg); c.alignment=CW; c.border=Bs('thin','FF93C5FD')
        hc(1,'No'); hc(2,'Name' if lang=='en' else 'Nama')
        for d in range(1,days_in+1):
            dt=datetime(yr2,mo2,d); dn=DNAMES[dt.weekday()]
            c=ws.cell(row=HDR,column=d+2,value=f'{d}\n{dn}')
            c.font=fnt(7,True,HT); c.alignment=CW; c.border=Bs('thin','FF93C5FD')
            if   dt.weekday()==6: c.fill=F('FF991B1B')
            elif dt.weekday()==5: c.fill=F('FF92400E')
            else:                 c.fill=F(CA)
        _lat_lbl='Late\n(min)' if lang=='en' else 'Terlambat\n(mnt)'
        _ot_lbl ='OT\n(min)'   if lang=='en' else 'Lembur\n(mnt)'
        _avg_lbl='Avg\nIn'     if lang=='en' else 'Rata\nMasuk'
        hc(TOT_COL,'∑\nHadir','FF065F46')
        hc(LAT_COL,_lat_lbl,'FF7C2D12')
        hc(OT_COL, _ot_lbl, 'FF1E40AF')
        hc(AVG_COL,_avg_lbl,'FF1E3A5F')

        for idx,name in enumerate(all_names,1):
            r=HDR+idx; ws.row_dimensions[r].height=17
            sub=mdata.filter_name(name)
            c=ws.cell(row=r,column=1,value=idx)
            c.font=fnt(8); c.fill=F(STR); c.alignment=C; c.border=Bs('thin',BD)
            c=ws.cell(row=r,column=2,value=name)
            c.font=fnt(9,True,'FF1E3A5F'); c.fill=F(STR); c.alignment=L; c.border=Bm('FF93C5FD')

            hadir=0; masuk_list=[]; total_late=0; total_ot=0
            for d in range(1,days_in+1):
                col=d+2; dt=datetime(yr2,mo2,d)
                dr=sub.filter_date(date(yr2,mo2,d))
                if len(dr)>0:
                    jam=dr._lst[0].jam_masuk; late=int(dr._lst[0].terlambat); ot=int(dr._lst[0].lembur)
                    c=ws.cell(row=r,column=col,value=jam)
                    masuk_list.append(jam); hadir+=1; total_late+=late; total_ot+=ot
                    if   late>0:          c.font=fnt(7,True,OTX); c.fill=F(OBG)
                    elif dt.weekday()==6: c.font=fnt(7,True,'FF7C3AED'); c.fill=F('FFEDE9FE')
                    elif dt.weekday()==5: c.font=fnt(7,True,'FF92400E'); c.fill=F(YBG)
                    else:                 c.font=fnt(7,False,GTX); c.fill=F(GBG)
                else:
                    c=ws.cell(row=r,column=col,value='')
                    if   dt.weekday()==6: c.fill=F('FFFCE7F3')
                    elif dt.weekday()==5: c.fill=F(YBG)
                    else:                 c.fill=F(RBG)
                    c.font=fnt(7)
                c.alignment=C; c.border=Bs('thin',BD)

            pct=round(hadir/days_in*100)
            c=ws.cell(row=r,column=TOT_COL,value=hadir)
            c.font=fnt(9,True,GTX if pct>=80 else RTX); c.fill=F(GBG if pct>=80 else RBG)
            c.alignment=C; c.border=Bm()
            c=ws.cell(row=r,column=LAT_COL,value=total_late if total_late else '-')
            c.font=fnt(9,total_late>0,OTX if total_late>0 else 'FF888888')
            c.fill=F(OBG if total_late>0 else R1); c.alignment=C; c.border=Bs('thin',BD)
            c=ws.cell(row=r,column=OT_COL,value=total_ot if total_ot else '-')
            c.font=fnt(9,total_ot>0,'FF1D4ED8' if total_ot>0 else 'FF888888')
            c.fill=F('FFE0EAFF' if total_ot>0 else R1); c.alignment=C; c.border=Bs('thin',BD)
            try:
                if masuk_list:
                    ts2=sum(datetime.strptime(t,'%H:%M').hour*3600+datetime.strptime(t,'%H:%M').minute*60 for t in masuk_list)
                    av=ts2//len(masuk_list); avg=f"{av//3600:02d}:{(av%3600)//60:02d}"
                else: avg='-'
            except: avg='-'
            c=ws.cell(row=r,column=AVG_COL,value=avg)
            c.font=fnt(8,False,'FF1E3A5F'); c.fill=F('FFE0EAFF'); c.alignment=C; c.border=Bs('thin',BD)

        # total harian
        rt=HDR+len(all_names)+1; ws.row_dimensions[rt].height=15
        ws.merge_cells(f'A{rt}:B{rt}')
        _tot_lbl='DAILY TOTAL' if lang=='en' else 'TOTAL HADIR HARIAN'
        c=ws.cell(row=rt,column=1,value=_tot_lbl)
        c.font=fnt(8,True,HT); c.fill=F(CA); c.alignment=C; c.border=Bs('thin',BD)
        for d in range(1,days_in+1):
            col=d+2; cnt=len(set(r.nama for r in mdata.filter_date(date(yr2,mo2,d))))
            c=ws.cell(row=rt,column=col,value=cnt if cnt else '')
            c.font=fnt(8,True,GTX if cnt else 'FFAAAAAA')
            c.fill=F(GBG if cnt else 'FFF1F5F9'); c.alignment=C; c.border=Bs('thin',BD)
        for col in (TOT_COL,LAT_COL,OT_COL,AVG_COL):
            c=ws.cell(row=rt,column=col,value=''); c.fill=F(CA); c.border=Bs('thin',BD)

        # legenda
        rl=rt+2; ws.row_dimensions[rl].height=13
        _legs = ([('On time',GBG,GTX),('Late',OBG,OTX),('Saturday',YBG,'FF92400E'),
                  ('Sunday','FFEDE9FE','FF7C3AED'),('Absent',RBG,RTX)]
                 if lang=='en' else
                 [('Tepat waktu',GBG,GTX),('Terlambat',OBG,OTX),('Sabtu',YBG,'FF92400E'),
                  ('Minggu','FFEDE9FE','FF7C3AED'),('Tidak hadir',RBG,RTX)])
        cl=1
        for lb,bg,tc in _legs:
            ws.merge_cells(start_row=rl,start_column=cl,end_row=rl,end_column=cl+2)
            c=ws.cell(row=rl,column=cl,value=f'■ {lb}')
            c.font=fnt(8,False,tc); c.fill=F(bg); c.alignment=L; cl+=3

    # ── REKAP ─────────────────────────────────────────────────────────────────
    wr=wb.create_sheet('Recap' if lang=='en' else 'Rekap'); wr.sheet_view.showGridLines=False
    for i,w in enumerate([5,15,15,8,8,10,8,12,10,10],1):
        wr.column_dimensions[get_column_letter(i)].width=w
    wr.merge_cells('A1:J1')
    _rh=('ATTENDANCE RECAP — CV REJEKI AMERTA JAYA' if lang=='en'
         else 'REKAP ABSENSI — CV REJEKI AMERTA JAYA')
    wr['A1']=_rh; wr['A1'].font=fnt(13,True,HT); wr['A1'].fill=F(CH); wr['A1'].alignment=C
    wr.row_dimensions[1].height=24
    wr.merge_cells('A2:J2')
    _tmin=daily.min_date().strftime('%d %B %Y'); _tmax=daily.max_date().strftime('%d %B %Y')
    _std_lbl='Std Check-in' if lang=='en' else 'Jam Masuk Std'
    _tol_lbl='Tolerance' if lang=='en' else 'Toleransi'
    wr['A2']=f"Period: {_tmin} - {_tmax}  |  {_std_lbl}: {jam_masuk_std}  |  {_tol_lbl}: {toleransi} min"
    wr['A2'].font=fnt(9,False,'FF444444',italic=True); wr['A2'].alignment=C
    wr['A2'].fill=F('FFF0F4FF'); wr.row_dimensions[2].height=15

    _rhdrs=(['No','Name','Month','Days','Present','Absent','% Present','Late (min)','OT (min)','Status']
            if lang=='en' else
            ['No','Nama','Bulan','Hari','Hadir','Tdk Hadir','% Hadir','Terlambat (mnt)','Lembur (mnt)','Status'])
    for i,h in enumerate(_rhdrs,1):
        c=wr.cell(row=3,column=i,value=h)
        c.font=fnt(9,True,HT); c.fill=F(CA); c.alignment=CW; c.border=Bs('thin','FF93C5FD')
    wr.row_dimensions[3].height=22
    r=4; no=1
    for name in all_names:
        sub=daily.filter_name(name)
        for month in months:
            yr2,mo2=int(month[:4]),int(month[5:])
            _,days_in=calendar.monthrange(yr2,mo2)
            md=sub.filter_month(month)
            if len(md)==0: continue
            hadir=len(md); tidak=days_in-hadir; pct=round(hadir/days_in*100)
            tl=sum(r2.terlambat for r2 in md); to=sum(r2.lembur for r2 in md)
            if lang=='en': status='Good' if pct>=90 else ('Fair' if pct>=75 else 'Poor')
            else:          status='Baik' if pct>=90 else ('Cukup' if pct>=75 else 'Kurang')
            bg=F(R1) if r%2==1 else F(R2)
            vals=[no,name,f"{MNAMES[mo2]} {yr2}",days_in,hadir,tidak,f"{pct}%",
                  tl if tl else '-',to if to else '-',status]
            for i,v in enumerate(vals,1):
                c=wr.cell(row=r,column=i,value=v)
                c.font=fnt(9,i==2); c.fill=bg; c.alignment=C if i!=2 else L; c.border=Bs('thin',BD)
            pc=wr.cell(row=r,column=7); sc=wr.cell(row=r,column=10)
            if pct>=90:
                for x in (pc,sc): x.font=fnt(9,True,GTX); x.fill=F(GBG)
            elif pct>=75:
                for x in (pc,sc): x.font=fnt(9,True,YTX); x.fill=F(YBG)
            else:
                for x in (pc,sc): x.font=fnt(9,True,RTX); x.fill=F(RBG)
            lc=wr.cell(row=r,column=8)
            if tl>0: lc.font=fnt(9,True,OTX); lc.fill=F(OBG)
            wr.row_dimensions[r].height=17; r+=1; no+=1

    # ── LOG DETAIL ────────────────────────────────────────────────────────────
    wd=wb.create_sheet('Log Detail'); wd.sheet_view.showGridLines=False
    for i, w in enumerate([5,14,12,10,10,10,8,10,10,16], 1):
        wd.column_dimensions[get_column_letter(i)].width = w
    wd.merge_cells('A1:J1')
    _lh=('ATTENDANCE LOG DETAIL — CV REJEKI AMERTA JAYA' if lang=='en'
         else 'LOG DETAIL ABSENSI — CV REJEKI AMERTA JAYA')
    wd['A1']=_lh; wd['A1'].font=fnt(12,True,HT); wd['A1'].fill=F(CH); wd['A1'].alignment=C
    wd.row_dimensions[1].height=22
    _lhdrs=(['No','Name','Date','Day','Check-in','Check-out','Taps','Late','OT','Flags']
            if lang=='en' else
            ['No','Nama','Tanggal','Hari','Jam Masuk','Jam Keluar','Tap','Terlambat','Lembur','Flag'])
    for i,h in enumerate(_lhdrs,1):
        c=wd.cell(row=2,column=i,value=h)
        c.font=fnt(9,True,HT); c.fill=F(CA); c.alignment=C; c.border=Bs('thin','FF93C5FD')
    wd.row_dimensions[2].height=18
    ds=daily.sort_by('tanggal','nama')
    for i,row in enumerate(ds):
        r3=i+3; tgl=row.tanggal
        dn=DFULL[tgl.weekday()]; bg=F(R1) if i%2==0 else F(R2)
        late=int(row.terlambat); ot=int(row.lembur)
        tap_info=f"{row.tap_total}x" if hasattr(row,'tap_total') else '-'
        flags = []
        if getattr(row, 'no_checkout', False):
            flags.append('NO_CHECKOUT' if lang == 'en' else 'TANPA_KELUAR')
        if getattr(row, 'recovered', False):
            flags.append('RECOVERED' if lang == 'en' else 'DIPULIHKAN')
        if getattr(row, 'pulang_cepat', 0):
            flags.append(f"EARLY-{row.pulang_cepat}m" if lang == 'en' else f"CEPAT-{row.pulang_cepat}m")
        flag_s = ', '.join(flags) if flags else '-'
        vals=[i+1,row.nama,tgl.strftime('%d/%m/%Y'),dn,
              row.jam_masuk,row.jam_keluar,tap_info,
              f"{late} min" if late>0 else '-',
              f"{ot} min"   if ot>0   else '-',
              flag_s]
        for col,v in enumerate(vals,1):
            c=wd.cell(row=r3,column=col,value=v)
            c.font=fnt(9,col==2)
            if col==8 and late>0: c.font=fnt(9,True,OTX); c.fill=F(OBG)
            elif col==9 and ot>0: c.font=fnt(9,True,'FF1D4ED8'); c.fill=F('FFE0EAFF')
            elif col==7 and hasattr(row,'tap_total') and row.tap_total>2:
                c.font=fnt(9,True,'FF6B21A8'); c.fill=F('FFEDE9FE')
            elif col==10 and flags:
                c.font=fnt(8,True,RTX); c.fill=F(RBG)
            elif getattr(row, 'recovered', False):
                c.fill = F('FFE0E7FF')
            else: c.fill=bg
            c.alignment=C if col!=2 else L; c.border=Bs('thin',BD)
        if getattr(row, 'no_checkout', False):
            # highlight checkout cell
            c6 = wd.cell(row=r3, column=6)
            c6.font = fnt(9, True, RTX); c6.fill = F(RBG)
        wd.row_dimensions[r3].height=15

    ms=[s for s in wb.sheetnames if s not in ('Recap','Rekap','Log Detail')]
    wb._sheets=[wb[s] for s in ms+[s for s in ('Recap','Rekap','Log Detail') if s in wb.sheetnames]]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_excel_for_preview(data_bytes, sheet_idx=0):
    """Parse Excel bytes and return (headers, rows) for Treeview display."""
    import io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
        ws = wb.worksheets[sheet_idx]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows: return [], []
        # skip merged title rows (usually rows 1-3), find header row
        # header = first row where most cells are non-None
        hdr_idx = 0
        for i, row in enumerate(all_rows):
            non_none = sum(1 for c in row if c is not None)
            if non_none >= 3:
                hdr_idx = i
                break
        headers = [str(c) if c is not None else '' for c in all_rows[hdr_idx]]
        # remove empty trailing cols
        while headers and not headers[-1]: headers.pop()
        n_cols = len(headers)
        data_rows = []
        for row in all_rows[hdr_idx+1:]:
            r = [str(c) if c is not None else '' for c in row[:n_cols]]
            if any(c for c in r): data_rows.append(r)
        wb.close()
        return headers, data_rows
    except Exception as e:
        return ['Error'], [[str(e)]]


# ─────────────────────────────────────────────────────────────────────────────
# UI — PySide6 (Qt). Logic layer above is UI-agnostic; everything below is view.
# ─────────────────────────────────────────────────────────────────────────────
ACCENT = '#1A56DB'

THEMES = {
    'light': dict(bg='#F1F5F9', card='#FFFFFF', border='#E2E8F0', input_border='#CBD5E1',
                  text='#0F172A', subtext='#475569', input_bg='#FFFFFF',
                  alt='#F8FAFC', sel='#DBEAFE', sel_text='#0F172A',
                  head='#F1F5F9', head_text='#334155', disabled='#94A3B8',
                  info='#1e40af'),
    'dark':  dict(bg='#0B1220', card='#1E293B', border='#334155', input_border='#475569',
                  text='#E2E8F0', subtext='#94A3B8', input_bg='#0F172A',
                  alt='#243247', sel='#1e40af', sel_text='#FFFFFF',
                  head='#16223A', head_text='#CBD5E1', disabled='#64748B',
                  info='#93C5FD'),
}

def build_qss(theme):
    t = THEMES[theme]
    return f"""
* {{ font-family: 'Segoe UI'; font-size: 9pt; }}
QMainWindow, QDialog {{ background: {t['bg']}; }}
QLabel, QCheckBox, QRadioButton {{ color: {t['text']}; font-weight: 400; background: transparent; }}
QCheckBox::indicator, QRadioButton::indicator {{
    width: 14px; height: 14px; background: {t['input_bg']};
    border: 2px solid {t['input_border']};
}}
QCheckBox::indicator {{ border-radius: 4px; }}
QRadioButton::indicator {{ border-radius: 9px; }}
QCheckBox::indicator:hover, QRadioButton::indicator:hover {{ border-color: {ACCENT}; }}
QCheckBox::indicator:checked, QRadioButton::indicator:checked {{
    background: {ACCENT}; border-color: {ACCENT};
}}
QWidget#hdr QLabel {{ color: white; }}
QWidget#hdr QPushButton {{
    background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.45); color: white;
}}
QWidget#hdr QPushButton:hover {{ background: white; color: {ACCENT}; }}
QGroupBox {{
    background: {t['card']}; border: 1px solid {t['border']}; border-radius: 10px;
    margin-top: 9px; padding: 10px 8px 8px 8px; font-weight: 600; color: {ACCENT};
}}
QGroupBox::title {{ subcontrol-origin: margin; left: 12px; padding: 0 4px; }}
QPushButton {{
    background: {t['card']}; border: 1px solid {t['input_border']}; border-radius: 6px;
    padding: 5px 8px; color: {t['text']};
}}
QPushButton:hover {{ background: {ACCENT}; border-color: {ACCENT}; color: white; }}
QPushButton:pressed {{ background: #1e40af; color: white; }}
QPushButton:disabled {{ background: {t['bg']}; color: {t['disabled']}; border-color: {t['border']}; }}
QPushButton[accent="true"] {{ background: {ACCENT}; border-color: {ACCENT}; color: white; font-weight: 600; }}
QPushButton[accent="true"]:hover {{ background: #1e40af; }}
QLineEdit, QComboBox {{
    background: {t['input_bg']}; border: 1px solid {t['input_border']}; border-radius: 6px;
    padding: 4px 8px; color: {t['text']};
}}
QLineEdit:focus, QComboBox:focus {{ border-color: {ACCENT}; }}
QComboBox::drop-down {{ border: none; width: 18px; }}
QComboBox QAbstractItemView {{ background: {t['card']}; color: {t['text']};
    selection-background-color: {ACCENT}; selection-color: white; }}
QTabWidget::pane {{ background: {t['card']}; border: 1px solid {t['border']}; border-radius: 10px; top: -1px; }}
QTabBar::tab {{
    background: transparent; color: {t['subtext']}; padding: 7px 16px; margin-right: 4px;
    border-top-left-radius: 8px; border-top-right-radius: 8px;
}}
QTabBar::tab:selected {{ background: {t['card']}; color: {ACCENT}; font-weight: 600;
    border: 1px solid {t['border']}; border-bottom: none; }}
QTabBar::tab:hover:!selected {{ color: {ACCENT}; }}
QTableWidget {{
    background: {t['card']}; border: none; gridline-color: {t['bg']};
    alternate-background-color: {t['alt']}; color: {t['text']};
    selection-background-color: {t['sel']}; selection-color: {t['sel_text']};
}}
QHeaderView::section {{
    background: {t['head']}; color: {t['head_text']}; font-weight: 600; border: none;
    border-bottom: 2px solid {t['border']}; padding: 6px 4px;
}}
QPlainTextEdit {{ background: #0F172A; color: #CBD5E1; border-radius: 8px;
    font-family: Consolas; font-size: 8pt; border: none; padding: 6px; }}
QSplitter::handle {{ background: {t['border']}; width: 3px; }}
QStatusBar {{ background: {t['head']}; color: {t['head_text']}; }}
QMessageBox QLabel {{ color: {t['text']}; }}
QMenu {{ background: {t['card']}; color: {t['text']}; border: 1px solid {t['border']}; }}
QMenu::item:selected {{ background: {ACCENT}; color: white; }}
QProgressBar {{ background: {t['border']}; border-radius: 6px; text-align: center;
    height: 14px; color: {t['text']}; }}
QProgressBar::chunk {{ background: {ACCENT}; border-radius: 6px; }}
"""

def _icon():
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app_icon.png')
    return QIcon(p) if os.path.exists(p) else QIcon()

def _fill_table(tbl, rows, row_colors=None):
    """rows = list of value-tuples; row_colors = list of hex or None per row."""
    tbl.setRowCount(len(rows))
    for r, vals in enumerate(rows):
        for c, v in enumerate(vals):
            it = QTableWidgetItem(str(v))
            it.setFlags(it.flags() & ~Qt.ItemIsEditable)
            if row_colors and row_colors[r]:
                it.setBackground(QColor(row_colors[r]))
            tbl.setItem(r, c, it)

def _mk_table(headers, widths=None, stretch_col=None):
    t = QTableWidget(0, len(headers))
    t.setHorizontalHeaderLabels(headers)
    t.verticalHeader().setVisible(False)
    t.setAlternatingRowColors(True)
    t.setSelectionBehavior(QAbstractItemView.SelectRows)
    t.setSelectionMode(QAbstractItemView.SingleSelection)
    t.setEditTriggers(QAbstractItemView.NoEditTriggers)
    t.verticalHeader().setDefaultSectionSize(26)
    if widths:
        for i, w in enumerate(widths):
            if w: t.setColumnWidth(i, w)
    if stretch_col is not None:
        t.horizontalHeader().setSectionResizeMode(stretch_col, QHeaderView.Stretch)
    return t


# ─────────────────────────────────────────────────────────────────────────────
# DIALOGS
# ─────────────────────────────────────────────────────────────────────────────
class SettingsDialog(QDialog):
    def __init__(self, parent, cfg, on_save):
        super().__init__(parent)
        self.setWindowTitle('Settings'); self.setModal(True)
        self.resize(520, 560)
        self.cfg = cfg; self.on_save = on_save
        lay = QVBoxLayout(self)
        form = QFormLayout()
        fields = [('IP Address','ip'),('Port','port'),
                  ('Comm password (device)','comm_password'),
                  ('Standard Check-in (HH:MM)','jam_masuk'),
                  ('Standard Check-out (HH:MM)','jam_keluar'),
                  ('Late Tolerance (minutes)','toleransi'),
                  ('Recovery anchor date (YYYY-MM-DD, blank=auto)','anomaly_anchor')]
        self.edits = {}
        for label, key in fields:
            e = QLineEdit(str(cfg.get(key,'')))
            form.addRow(label, e); self.edits[key] = e
        lay.addLayout(form)
        # punch mode
        prow = QHBoxLayout()
        prow.addWidget(QLabel('In/Out mode:'))
        self.punch_cb = QComboBox()
        self.punch_cb.addItem('First/Last tap (face verify)', 'first_last')
        self.punch_cb.addItem('Forced device punch status', 'device_punch')
        pm = cfg.get('punch_mode', 'first_last')
        self.punch_cb.setCurrentIndex(0 if pm != 'device_punch' else 1)
        prow.addWidget(self.punch_cb); prow.addStretch()
        lay.addLayout(prow)
        note = QLabel('Forced mode: uses punch 0/4 = in, 1/5 = out from device.\n'
                      'If multi-day gaps found during recovery, set anchor date manually.')
        note.setStyleSheet('color:#888; font-size:8pt;'); lay.addWidget(note)
        self.checks = {}
        for label, key, dflt in [
                ('Auto backup CSV after pull','auto_backup',False),
                ('Auto-recover clock-reset (year 2000) records','anomaly_recover',True),
                ('Auto-start Live Monitor + punch notifications','live_autostart',False)]:
            cb = QCheckBox(label); cb.setChecked(bool(cfg.get(key,dflt)))
            lay.addWidget(cb); self.checks[key] = cb

        # Silent mode / autostart
        silent_box = QGroupBox('Silent Mode (Tray + Auto Pull)')
        sl = QVBoxLayout(silent_box)
        self.checks['silent_mode'] = QCheckBox(
            'Silent mode — jalankan di system tray (tanpa jendela)')
        self.checks['silent_mode'].setChecked(bool(cfg.get('silent_mode', False)))
        sl.addWidget(self.checks['silent_mode'])
        self.checks['autostart'] = QCheckBox(
            'Start with Windows (login → tray + auto pull)')
        self.checks['autostart'].setChecked(bool(
            cfg.get('autostart', False) or cfg.get('silent_mode', False)))
        sl.addWidget(self.checks['autostart'])
        self.checks['auto_pull_on_start'] = QCheckBox(
            'Auto Pull + Sync ke cloud saat PC/app start')
        self.checks['auto_pull_on_start'].setChecked(bool(cfg.get('auto_pull_on_start', True)))
        sl.addWidget(self.checks['auto_pull_on_start'])
        row_iv = QHBoxLayout()
        row_iv.addWidget(QLabel('Ulangi auto-pull tiap (menit, 0=hanya sekali):'))
        self.edits['auto_pull_interval_min'] = QLineEdit(str(cfg.get('auto_pull_interval_min', 60)))
        self.edits['auto_pull_interval_min'].setFixedWidth(60)
        row_iv.addWidget(self.edits['auto_pull_interval_min'])
        row_iv.addStretch()
        sl.addLayout(row_iv)
        st = QLabel(
            'Saat Windows nyala: app ke tray → tunggu jaringan → Pull mesin → '
            'sync VST (jika cloud sync aktif). Tutup jendela = tetap di tray.'
        )
        st.setStyleSheet('color:#888; font-size:8pt;')
        st.setWordWrap(True)
        sl.addWidget(st)
        lay.addWidget(silent_box)

        # Cloud Sync → VST-laravel
        cloud_box = QGroupBox('Cloud Sync (VST Absensi)')
        cl = QVBoxLayout(cloud_box)
        self.checks['cloud_sync_enabled'] = QCheckBox('Enable sync to VST-laravel')
        self.checks['cloud_sync_enabled'].setChecked(bool(cfg.get('cloud_sync_enabled', False)))
        cl.addWidget(self.checks['cloud_sync_enabled'])
        self.checks['cloud_sync_after_pull'] = QCheckBox('Auto-sync after every Pull')
        self.checks['cloud_sync_after_pull'].setChecked(bool(cfg.get('cloud_sync_after_pull', True)))
        cl.addWidget(self.checks['cloud_sync_after_pull'])
        cf = QFormLayout()
        self.edits['cloud_api_url'] = QLineEdit(str(cfg.get(
            'cloud_api_url', 'https://service.rejekiamerta.com/api/attendance')))
        self.edits['cloud_api_token'] = QLineEdit(str(cfg.get('cloud_api_token', '')))
        self.edits['cloud_api_token'].setEchoMode(QLineEdit.Password)
        self.edits['cloud_api_token'].setPlaceholderText(
            'Token dari VST: Pengaturan → Sync ZKTeco Desktop')
        # toggle show token (Password mode sering bikin paste salah / susah cek)
        self._tok_visible = False
        tok_row = QHBoxLayout()
        tok_row.addWidget(self.edits['cloud_api_token'], 1)
        self._btn_show_tok = QPushButton('👁')
        self._btn_show_tok.setFixedWidth(36)
        self._btn_show_tok.setToolTip('Tampilkan / sembunyikan token')
        self._btn_show_tok.clicked.connect(self._toggle_token_visible)
        tok_row.addWidget(self._btn_show_tok)
        cf.addRow('API Base URL', self.edits['cloud_api_url'])
        cf.addRow('API Token', tok_row)
        cl.addLayout(cf)
        tip = QLabel(
            '1) Login VST → Pengaturan → Sync ZKTeco Desktop\n'
            '2) Salin API Base URL + Token (hanya hex, tanpa "Bearer ")\n'
            '3) Setelah Generate Ulang Token di web, update token di sini\n'
            'URL default: https://service.rejekiamerta.com/api/attendance'
        )
        tip.setStyleSheet('color:#888; font-size:8pt;')
        tip.setWordWrap(True)
        cl.addWidget(tip)
        test_btn = QPushButton('🔗 Tes koneksi VST')
        test_btn.clicked.connect(self._test_cloud)
        cl.addWidget(test_btn)
        lay.addWidget(cloud_box)

        btns = QHBoxLayout()
        ok = QPushButton('💾 Save'); ok.setProperty('accent', True); ok.clicked.connect(self._save)
        cancel = QPushButton('Cancel'); cancel.clicked.connect(self.reject)
        btns.addStretch(); btns.addWidget(ok); btns.addWidget(cancel)
        lay.addLayout(btns)

    def _toggle_token_visible(self):
        self._tok_visible = not self._tok_visible
        mode = QLineEdit.Normal if self._tok_visible else QLineEdit.Password
        self.edits['cloud_api_token'].setEchoMode(mode)

    def _collect_cfg(self):
        for key, e in self.edits.items():
            v = e.text().strip()
            if key in ('toleransi', 'comm_password', 'port', 'auto_pull_interval_min'):
                try:
                    v = int(v)
                except ValueError:
                    pass
            self.cfg[key] = v
        self.cfg['punch_mode'] = self.punch_cb.currentData()
        for key, cb in self.checks.items():
            self.cfg[key] = cb.isChecked()
        if self.cfg.get('silent_mode'):
            self.cfg['autostart'] = True
        # normalisasi credential agar paste "Bearer xxx" / URL .../sync tidak gagal
        self.cfg['cloud_api_url'] = normalize_cloud_base_url(self.cfg.get('cloud_api_url') or '')
        self.cfg['cloud_api_token'] = normalize_cloud_token(self.cfg.get('cloud_api_token') or '')
        return self.cfg

    def _test_cloud(self):
        cfg = dict(self._collect_cfg())
        # pastikan tes jalan meski checkbox belum dicentang
        cfg['cloud_sync_enabled'] = True
        try:
            result = cloud_credentials_ok(cfg)
            QMessageBox.information(self, 'Tes VST', result.get('message', 'OK'))
        except Exception as e:
            QMessageBox.warning(self, 'Tes VST gagal', str(e))

    def _save(self):
        self._collect_cfg()
        # sync field UI ke nilai yang sudah dinormalisasi
        self.edits['cloud_api_url'].setText(self.cfg.get('cloud_api_url') or '')
        self.edits['cloud_api_token'].setText(self.cfg.get('cloud_api_token') or '')
        save_config(self.cfg)
        self.on_save(self.cfg)
        self.accept()


class DeviceInfoDialog(QDialog):
    def __init__(self, parent, info_dict):
        super().__init__(parent)
        self.setWindowTitle('Device Info'); self.setModal(True)
        lay = QVBoxLayout(self)
        hdr = QLabel('Device Info — eFace10')
        hdr.setStyleSheet(f'background:{ACCENT}; color:white; font-weight:600; '
                          'padding:8px 12px; border-radius:6px;')
        lay.addWidget(hdr)
        form = QFormLayout()
        for k, v in info_dict.items():
            lbl = QLabel(f'<b>{k}:</b>'); val = QLabel(str(v))
            form.addRow(lbl, val)
        lay.addLayout(form)
        close = QPushButton('Close'); close.clicked.connect(self.accept)
        lay.addWidget(close, alignment=Qt.AlignCenter)


class UserManagerDialog(QDialog):
    def __init__(self, parent, users, app=None):
        super().__init__(parent)
        self.app = app
        self.setWindowTitle('Users on Device'); self.setModal(True)
        self.resize(460, 480)
        lay = QVBoxLayout(self)
        hdr = QLabel('Users registered on eFace10')
        hdr.setStyleSheet(f'background:{ACCENT}; color:white; font-weight:600; '
                          'padding:8px 12px; border-radius:6px;')
        lay.addWidget(hdr)
        self.table = _mk_table(['UID','Name','Card ID'], [70,180,120], stretch_col=1)
        self.table.itemSelectionChanged.connect(self._on_select)
        lay.addWidget(self.table)
        ef = QHBoxLayout()
        ef.addWidget(QLabel('UID:'))
        self.uid_edit = QLineEdit(); self.uid_edit.setFixedWidth(60); ef.addWidget(self.uid_edit)
        ef.addWidget(QLabel('Name:'))
        self.name_edit = QLineEdit(); ef.addWidget(self.name_edit)
        save = QPushButton('💾 Add / Rename'); save.clicked.connect(self._save_user); ef.addWidget(save)
        dele = QPushButton('🗑 Delete'); dele.clicked.connect(self._delete_user); ef.addWidget(dele)
        lay.addLayout(ef)
        note = QLabel('Face/fingerprint enrollment is done on the device itself.')
        note.setStyleSheet('color:#888; font-size:8pt;'); lay.addWidget(note)
        bf = QHBoxLayout()
        self.total_lbl = QLabel(); bf.addWidget(self.total_lbl); bf.addStretch()
        close = QPushButton('Close'); close.clicked.connect(self.accept); bf.addWidget(close)
        lay.addLayout(bf)
        self._fill(users)

    def _fill(self, users):
        _fill_table(self.table, [(u['uid'], u['nama'], u.get('card_id','')) for u in users])
        self.total_lbl.setText(f'Total: {len(users)} users')

    def _on_select(self):
        r = self.table.currentRow()
        if r >= 0:
            self.uid_edit.setText(self.table.item(r,0).text())
            self.name_edit.setText(self.table.item(r,1).text())

    def _save_user(self):
        uid = self.uid_edit.text().strip(); name = self.name_edit.text().strip()
        if not uid.isdigit() or not name:
            QMessageBox.warning(self,'Invalid','UID must be a number and name cannot be empty.'); return
        if self.app: self.app.device_user_save(uid, name, self)

    def _delete_user(self):
        uid = self.uid_edit.text().strip()
        if not uid.isdigit():
            QMessageBox.warning(self,'Invalid','Select a user or enter a UID first.'); return
        if QMessageBox.question(self,'Delete',
                f'Delete user {uid} from the DEVICE?\n'
                'Face/fingerprint templates on the device are removed too.') != QMessageBox.Yes:
            return
        if self.app: self.app.device_user_delete(uid, self)


class LeaveDialog(QDialog):
    """Manage izin / cuti / dinas / sakit per employee-day (local SQLite)."""
    def __init__(self, parent, cfg):
        super().__init__(parent)
        self.cfg = cfg
        self.setWindowTitle('Izin / Cuti / Dinas / Sakit')
        self.setModal(True)
        self.resize(520, 420)
        lay = QVBoxLayout(self)
        form = QHBoxLayout()
        form.addWidget(QLabel('UID:'))
        self.uid_edit = QLineEdit(); self.uid_edit.setFixedWidth(50); form.addWidget(self.uid_edit)
        form.addWidget(QLabel('Tanggal (YYYY-MM-DD):'))
        self.date_edit = QLineEdit(date.today().strftime('%Y-%m-%d')); form.addWidget(self.date_edit)
        form.addWidget(QLabel('Jenis:'))
        self.jenis_cb = QComboBox(); self.jenis_cb.addItems(list(LEAVE_TYPES)); form.addWidget(self.jenis_cb)
        form.addWidget(QLabel('Catatan:'))
        self.note_edit = QLineEdit(); form.addWidget(self.note_edit)
        b_add = QPushButton('➕ Simpan'); b_add.clicked.connect(self._add); form.addWidget(b_add)
        lay.addLayout(form)
        self.table = _mk_table(['UID', 'Nama', 'Tanggal', 'Jenis', 'Catatan'], [50, 120, 100, 70, 140], stretch_col=4)
        lay.addWidget(self.table, 1)
        row = QHBoxLayout()
        b_del = QPushButton('🗑 Hapus baris terpilih'); b_del.clicked.connect(self._del)
        b_close = QPushButton('Close'); b_close.clicked.connect(self.accept)
        row.addWidget(b_del); row.addStretch(); row.addWidget(b_close)
        lay.addLayout(row)
        self._reload()

    def _reload(self):
        um = {int(k): v for k, v in self.cfg.get('user_map', {}).items()}
        rows = db_list_leaves()
        _fill_table(self.table, [
            (r['uid'], um.get(r['uid'], f"UID:{r['uid']}"), r['tanggal'], r['jenis'], r['note'])
            for r in rows
        ])

    def _add(self):
        uid = self.uid_edit.text().strip()
        tgl = self.date_edit.text().strip()
        if not uid.isdigit():
            QMessageBox.warning(self, 'Invalid', 'UID harus angka.'); return
        try:
            datetime.strptime(tgl, '%Y-%m-%d')
        except ValueError:
            QMessageBox.warning(self, 'Invalid', 'Tanggal harus YYYY-MM-DD.'); return
        db_set_leave(int(uid), tgl, self.jenis_cb.currentText(), self.note_edit.text().strip())
        self._reload()

    def _del(self):
        r = self.table.currentRow()
        if r < 0: return
        uid = int(self.table.item(r, 0).text())
        tgl = self.table.item(r, 2).text()
        db_delete_leave(uid, tgl)
        self._reload()


class HolidayDialog(QDialog):
    """Set company-wide holiday dates (manual). Weekend is NOT auto-libur."""
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle('Hari Libur (Manual)')
        self.setModal(True)
        self.resize(480, 400)
        lay = QVBoxLayout(self)
        tip = QLabel(
            'Sabtu/Minggu TIDAK otomatis libur.\n'
            'Set tanggal libur event (nasional / cuti bersama / libur kantor) di bawah.'
        )
        tip.setStyleSheet('color:#555; font-size:9pt;')
        tip.setWordWrap(True)
        lay.addWidget(tip)

        form = QHBoxLayout()
        form.addWidget(QLabel('Tanggal (YYYY-MM-DD):'))
        self.date_edit = QLineEdit(date.today().strftime('%Y-%m-%d'))
        self.date_edit.setFixedWidth(110)
        form.addWidget(self.date_edit)
        form.addWidget(QLabel('Keterangan:'))
        self.note_edit = QLineEdit()
        self.note_edit.setPlaceholderText('mis. Idul Fitri, cuti bersama…')
        form.addWidget(self.note_edit, 1)
        lay.addLayout(form)

        self.confirm_cb = QCheckBox(
            'Saya konfirmasi tanggal ini adalah HARI LIBUR untuk semua karyawan'
        )
        lay.addWidget(self.confirm_cb)

        b_add = QPushButton('➕ Tetapkan sebagai Hari Libur')
        b_add.setProperty('accent', True)
        b_add.clicked.connect(self._add)
        lay.addWidget(b_add)

        self.table = _mk_table(['Tanggal', 'Hari', 'Keterangan'], [110, 80, 220], stretch_col=2)
        lay.addWidget(self.table, 1)

        row = QHBoxLayout()
        b_del = QPushButton('🗑 Hapus libur terpilih')
        b_del.clicked.connect(self._del)
        b_close = QPushButton('Close')
        b_close.clicked.connect(self.accept)
        row.addWidget(b_del)
        row.addStretch()
        row.addWidget(b_close)
        lay.addLayout(row)
        self._reload()

    def _reload(self):
        rows = db_list_holidays()
        data = []
        for r in rows:
            try:
                d = datetime.strptime(r['tanggal'], '%Y-%m-%d').date()
                hari = HARI_ID[d.weekday()]
            except Exception:
                hari = ''
            data.append((r['tanggal'], hari, r['note']))
        _fill_table(self.table, data)

    def _add(self):
        tgl = self.date_edit.text().strip()
        try:
            d = datetime.strptime(tgl, '%Y-%m-%d').date()
        except ValueError:
            QMessageBox.warning(self, 'Invalid', 'Tanggal harus YYYY-MM-DD.')
            return
        if not self.confirm_cb.isChecked():
            QMessageBox.warning(
                self, 'Konfirmasi',
                'Centang checklist konfirmasi dulu sebelum menetapkan hari libur.'
            )
            return
        note = self.note_edit.text().strip()
        if QMessageBox.question(
            self, 'Konfirmasi Hari Libur',
            f'Tetapkan {tgl} ({HARI_ID[d.weekday()]}) sebagai HARI LIBUR?\n\n'
            f'Keterangan: {note or "(kosong)"}\n\n'
            f'Semua karyawan tanpa izin/cuti akan berstatus LIBUR (bukan ABSEN).'
        ) != QMessageBox.Yes:
            return
        db_set_holiday(tgl, note)
        self.confirm_cb.setChecked(False)
        self.note_edit.clear()
        self._reload()
        QMessageBox.information(self, 'OK', f'{tgl} diset sebagai hari libur.')

    def _del(self):
        r = self.table.currentRow()
        if r < 0:
            return
        tgl = self.table.item(r, 0).text()
        if QMessageBox.question(
            self, 'Hapus', f'Hapus {tgl} dari daftar hari libur?'
        ) != QMessageBox.Yes:
            return
        db_delete_holiday(tgl)
        self._reload()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN WINDOW
# ─────────────────────────────────────────────────────────────────────────────
class _Bridge(QObject):
    call = Signal(object)   # thread-safe "run this on the UI thread"


class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f'ZKTeco eFace10 Utility v{APP_VERSION} — CV RAJ')
        self.setMinimumSize(1100, 680)
        self.setWindowIcon(_icon())
        self._bridge = _Bridge()
        self._bridge.call.connect(lambda f: f())
        self.ui = self._bridge.call.emit   # usable from any thread
        self.cfg = load_config()
        self._cache = []
        self._current_snap_bytes = None   # Excel bytes in memory
        self._current_snap_id = None
        self._toasts = []
        init_db()
        self._build_ui()
        self._apply_theme(self.cfg.get('theme', 'light'))
        self._update_status()
        self._refresh_history()
        self._refresh_today()
        try:
            from updater import cleanup_old_exe
            cleanup_old_exe()
        except ImportError: pass
        self._boot_silent = (
            '--silent' in sys.argv
            or '--minimized' in sys.argv
            or bool(self.cfg.get('silent_mode', False) and '--show' not in sys.argv)
        )
        # only force silent hide when launched with flag (Windows autostart)
        self._boot_silent_flag = '--silent' in sys.argv or '--minimized' in sys.argv
        self._tray_setup()
        if self.cfg.get('live_autostart', False):
            QTimer.singleShot(1500, self._toggle_live)
        self._clock_timer = QTimer(self)
        self._clock_timer.timeout.connect(self._clock_tick)
        self._clock_timer.start(600_000)   # background clock guard, every 10 min
        self._setup_auto_pull_schedule()

    # ── System tray ───────────────────────────────────────────────────────────
    def _tray_setup(self):
        self._tray = None
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return
        self._tray = QSystemTrayIcon(_icon(), self)
        self._tray.setToolTip(f'ZKTeco Utility v{APP_VERSION}')
        menu = QMenu()
        a_open = QAction('Open Dashboard', menu)
        a_open.triggered.connect(self._tray_open)
        a_pull = QAction('Pull + Sync Now', menu)
        a_pull.triggered.connect(lambda: self._run_quiet(lambda: self._do_pull(quiet=True)))
        a_sync = QAction('Sync to VST Cloud', menu)
        a_sync.triggered.connect(lambda: self._run_quiet(self._do_cloud_sync_quiet))
        a_exit = QAction('Exit', menu)
        a_exit.triggered.connect(self._tray_exit)
        menu.addAction(a_open)
        menu.addSeparator()
        menu.addAction(a_pull)
        menu.addAction(a_sync)
        menu.addSeparator()
        menu.addAction(a_exit)
        self._tray.setContextMenu(menu)
        # Double-click saja yang buka window — single click di tray sering
        # tidak sengaja (icon "jiggle"/overflow) dan terasa seperti glitch.
        self._tray.activated.connect(self._on_tray_activated)
        self._tray.show()

    def _on_tray_activated(self, reason):
        if reason == QSystemTrayIcon.DoubleClick:
            self._tray_open()

    def closeEvent(self, ev):
        # always prefer tray when available (silent mode)
        if self._tray:
            ev.ignore()
            self.hide()
            if self.cfg.get('silent_mode') or self._boot_silent_flag:
                self._tray_msg(
                    'ZKTeco Utility',
                    'Berjalan di tray. Double-click ikon untuk buka · klik kanan → Exit.',
                    QSystemTrayIcon.Information,
                )
        else:
            ev.accept()

    def _tray_open(self):
        self.showNormal()
        self.raise_()
        self.activateWindow()
        try:
            self.setWindowState(self.windowState() & ~Qt.WindowMinimized)
        except Exception:
            pass

    def _tray_exit(self):
        self._live_want = False
        if getattr(self, '_pull_timer', None):
            self._pull_timer.stop()
        if self._tray:
            self._tray.hide()
        QApplication.quit()

    def _tray_msg(self, title, msg, icon=None):
        if icon is None:
            icon = QSystemTrayIcon.Information
        if self._tray:
            try:
                self._tray.showMessage(title, msg, icon, 6000)
            except Exception:
                pass

    def _setup_auto_pull_schedule(self):
        """Boot auto-pull + optional interval while running in tray."""
        self._pull_timer = None
        if self.cfg.get('auto_pull_on_start', True) and self._boot_silent_flag:
            # delay: wait NIC/device after Windows login
            delay_ms = 12_000
            self._log(f'Silent: auto pull+sync in {delay_ms // 1000}s ...')
            QTimer.singleShot(delay_ms, self._startup_auto_pull)
            if self._tray:
                QTimer.singleShot(
                    500,
                    lambda: self._tray_msg(
                        'ZKTeco Silent',
                        'Siap di tray. Auto pull sebentar lagi…',
                    ),
                )
        try:
            mins = int(self.cfg.get('auto_pull_interval_min', 0) or 0)
        except (TypeError, ValueError):
            mins = 0
        if mins > 0:
            self._pull_timer = QTimer(self)
            self._pull_timer.timeout.connect(self._startup_auto_pull)
            self._pull_timer.start(mins * 60 * 1000)
            self._log(f'Auto-pull interval: setiap {mins} menit')

    def _startup_auto_pull(self):
        self._log('▶ Auto pull (silent) ...')
        self._run_quiet(lambda: self._do_pull(quiet=True))

    def _do_cloud_sync_quiet(self):
        try:
            cloud_sync(self.cfg, punches=None, log=self._log)
            self.ui(lambda: self._tray_msg('Cloud Sync', 'Sync ke VST berhasil.'))
        except Exception as e:
            self._log(f'⚠ Cloud sync gagal: {e}')
            self.ui(lambda: self._tray_msg('Cloud Sync', str(e), QSystemTrayIcon.Warning))

    def _clock_tick(self):
        """Every 10 min: make sure the device clock matches the PC (power-loss guard).
        # ponytail: skipped while live monitor runs — its reconnect cycle checks instead"""
        def check():
            conn = None
            try:
                conn = self._get_conn()
                self._check_clock(conn, quiet=True)
            except Exception: pass   # device off/unreachable — silent, retry next tick
            finally:
                try:
                    if conn: conn.disconnect()
                except Exception: pass
        if not getattr(self, '_live_want', False):
            threading.Thread(target=check, daemon=True).start()

    # ── UI construction ───────────────────────────────────────────────────────
    def _build_ui(self):
        central = QWidget(); self.setCentralWidget(central)
        root = QVBoxLayout(central); root.setContentsMargins(0,0,0,0); root.setSpacing(0)

        # Header
        hdr = QWidget(); hdr.setObjectName('hdr'); hdr.setStyleSheet(f'background:{ACCENT};')
        hl = QHBoxLayout(hdr); hl.setContentsMargins(14,8,14,8)
        title = QLabel('ZKTeco eFace10 Utility  ·  CV RAJ')
        title.setStyleSheet('color:white; font-size:12pt; font-weight:700;')
        hl.addWidget(title); hl.addStretch()
        self.lang_cb = QComboBox(); self.lang_cb.addItems(['en','id'])
        self.lang_cb.setCurrentText(self.cfg.get('lang','en'))
        self.lang_cb.currentTextChanged.connect(self._on_lang_change)
        self.lang_cb.setFixedWidth(56)
        hl.addWidget(QLabel('🌐')); hl.addWidget(self.lang_cb)
        self.theme_btn = QPushButton('🌙')
        self.theme_btn.setFixedWidth(36); self.theme_btn.setToolTip('Dark / light mode')
        self.theme_btn.clicked.connect(self._toggle_theme)
        hl.addWidget(self.theme_btn)
        ver = QLabel(f'v{APP_VERSION}'); ver.setStyleSheet('color:#93C5FD;')
        hl.addWidget(ver)
        b_upd = QPushButton('⬆ Update'); b_upd.clicked.connect(self._check_update); hl.addWidget(b_upd)
        b_set = QPushButton('⚙ Settings'); b_set.clicked.connect(self._open_settings); hl.addWidget(b_set)
        root.addWidget(hdr)

        split = QSplitter(Qt.Horizontal)
        root.addWidget(split, 1)

        # ══ LEFT PANEL ════════════════════════════════════════════════════════
        left = QWidget(); left.setMinimumWidth(380); left.setMaximumWidth(460)
        ll = QVBoxLayout(left); ll.setContentsMargins(8,8,4,8); ll.setSpacing(8)

        # Connection card
        gc = QGroupBox('Device Connection')
        gl = QGridLayout(gc)
        gl.addWidget(QLabel('IP:'), 0, 0)
        self.ip_edit = QLineEdit(self.cfg['ip']); gl.addWidget(self.ip_edit, 0, 1)
        gl.addWidget(QLabel('Port:'), 0, 2)
        self.port_edit = QLineEdit(str(self.cfg['port'])); self.port_edit.setFixedWidth(64)
        gl.addWidget(self.port_edit, 0, 3)
        self.conn_lbl = QLabel('● Not connected'); self.conn_lbl.setStyleSheet('color:#888; font-size:8pt;')
        gl.addWidget(self.conn_lbl, 0, 4)
        btns = [('🔌 Test Connection', lambda: self._run(self._do_test)),
                ('ℹ Device Info',      lambda: self._run(self._do_info)),
                ('📡 Live Monitor',    self._toggle_live),
                ('👤 Manage Users',    lambda: self._run(self._do_users)),
                ('♻ Restart Device',   self._confirm_restart)]
        for i, (txt, cmd) in enumerate(btns):   # 2 columns so labels never clip at 380px
            b = QPushButton(txt); b.clicked.connect(cmd)
            gl.addWidget(b, 1 + i//2, (i%2)*2, 1, 2)
            if txt.startswith('📡'): self.btn_live = b
        ll.addWidget(gc)

        # Workflow card
        gw = QGroupBox('Workflow')
        wl = QGridLayout(gw)
        self.btn_pull = QPushButton('📥 1 · Pull Data'); self.btn_pull.setProperty('accent', True)
        self.btn_pull.clicked.connect(lambda: self._run(self._do_pull))
        wl.addWidget(self.btn_pull, 0, 0, 1, 2)
        wl.addWidget(QLabel('Month:'), 1, 0)
        self.bulan_cb = QComboBox()
        self.bulan_cb.addItems(['All','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'])
        wl.addWidget(self.bulan_cb, 1, 1)
        wl.addWidget(QLabel('Year:'), 2, 0)
        now = datetime.now()
        self.tahun_cb = QComboBox()
        self.tahun_cb.addItems([str(y) for y in range(now.year-3, now.year+2)])
        self.tahun_cb.setCurrentText(str(now.year))
        wl.addWidget(self.tahun_cb, 2, 1)
        srcrow = QHBoxLayout()
        srcrow.addWidget(QLabel('Source:'))
        self.src_db = QRadioButton('Database'); self.src_db.setChecked(True)
        self.src_cache = QRadioButton('Cache')
        srcrow.addWidget(self.src_db); srcrow.addWidget(self.src_cache); srcrow.addStretch()
        wl.addLayout(srcrow, 3, 0, 1, 2)
        self.btn_report = QPushButton('📊 2 · Preview Report'); self.btn_report.setProperty('accent', True)
        self.btn_report.clicked.connect(lambda: self._run(self._do_report))
        wl.addWidget(self.btn_report, 4, 0, 1, 2)
        sc = QHBoxLayout()
        self.btn_all = QPushButton('⚡ All at Once'); self.btn_all.clicked.connect(lambda: self._run(self._do_all))
        b_clear = QPushButton('🗑 Clear Device Log'); b_clear.clicked.connect(self._confirm_clear)
        sc.addWidget(self.btn_all); sc.addWidget(b_clear)
        wl.addLayout(sc, 5, 0, 1, 2)
        sc2 = QHBoxLayout()
        b_leave = QPushButton('📋 Izin/Cuti/Dinas/Sakit'); b_leave.clicked.connect(self._open_leave)
        b_pay = QPushButton('💰 Export Payroll CSV'); b_pay.clicked.connect(self._export_payroll)
        sc2.addWidget(b_leave); sc2.addWidget(b_pay)
        wl.addLayout(sc2, 6, 0, 1, 2)
        b_hol = QPushButton('🏖 Hari Libur (Manual)'); b_hol.clicked.connect(self._open_holiday)
        wl.addWidget(b_hol, 7, 0, 1, 2)
        b_cloud = QPushButton('☁ Sync to VST Cloud'); b_cloud.clicked.connect(lambda: self._run(self._do_cloud_sync))
        wl.addWidget(b_cloud, 8, 0, 1, 2)
        self.data_lbl = QLabel('...'); self.data_lbl.setStyleSheet('color:#1e40af; font-size:8pt;')
        wl.addWidget(self.data_lbl, 9, 0, 1, 2, alignment=Qt.AlignCenter)
        ll.addWidget(gw)

        # Log card
        glog = QGroupBox('Log')
        loglay = QVBoxLayout(glog)
        self.log_box = QPlainTextEdit(); self.log_box.setReadOnly(True)
        self.log_box.setMaximumBlockCount(2000)
        loglay.addWidget(self.log_box)
        ll.addWidget(glog, 1)
        split.addWidget(left)

        # ══ RIGHT PANEL — tabs ════════════════════════════════════════════════
        right = QWidget()
        rl = QVBoxLayout(right); rl.setContentsMargins(4,8,8,8)
        self.tabs = QTabWidget()
        rl.addWidget(self.tabs)
        split.addWidget(right)
        split.setStretchFactor(1, 1)

        # Tab 0: Today
        tab_today = QWidget(); tl = QVBoxLayout(tab_today)
        trow = QHBoxLayout()
        cap = QLabel('Absensi hari ini'); cap.setStyleSheet('font-weight:600;')
        trow.addWidget(cap); trow.addStretch()
        b_ref = QPushButton('🔄 Refresh'); b_ref.clicked.connect(self._refresh_today)
        trow.addWidget(b_ref)
        tl.addLayout(trow)
        tiles = QHBoxLayout()
        self.tile_hadir = self._mktile(tiles, 'Hadir', '#D1FAE5', '#065F46')
        self.tile_telat = self._mktile(tiles, 'Telat', '#FED7AA', '#9A3412')
        self.tile_absen = self._mktile(tiles, 'Belum Absen', '#FEE2E2', '#991B1B')
        tl.addLayout(tiles)
        self.today_tbl = _mk_table(['Nama','Jam Masuk','Status'], [180,100,110], stretch_col=0)
        tl.addWidget(self.today_tbl, 1)
        self.tabs.addTab(tab_today, '🏠 Today')
        self.tabs.currentChanged.connect(lambda i: self._refresh_today() if i == 0 else None)

        # Tab 1: Report Viewer
        tab_view = QWidget(); vl = QVBoxLayout(tab_view)
        vt = QHBoxLayout()
        vt.addWidget(QLabel('Sheet:'))
        self.sheet_cb = QComboBox(); self.sheet_cb.setMinimumWidth(160)
        self.sheet_cb.activated.connect(self._on_sheet_change)
        vt.addWidget(self.sheet_cb)
        b_save = QPushButton('💾 Save to File'); b_save.clicked.connect(self._save_excel); vt.addWidget(b_save)
        b_rref = QPushButton('🔄 Refresh'); b_rref.clicked.connect(lambda: self._run(self._do_report)); vt.addWidget(b_rref)
        vt.addStretch()
        self.snap_lbl = QLabel('No report loaded'); self.snap_lbl.setStyleSheet('color:#64748B; font-size:8pt;')
        vt.addWidget(self.snap_lbl)
        vl.addLayout(vt)
        self.report_tbl = _mk_table([])
        vl.addWidget(self.report_tbl, 1)
        self.tabs.addTab(tab_view, '📊 Report Viewer')

        # Tab 2: Pull History
        tab_hist = QWidget(); hlv = QVBoxLayout(tab_hist)
        gh1 = QGroupBox('Pull Sessions'); g1 = QVBoxLayout(gh1)
        self.sess_tbl = _mk_table(['ID','Pull Date','Records','New','Device IP'], [46,160,80,80,120], stretch_col=1)
        g1.addWidget(self.sess_tbl)
        sb = QHBoxLayout()
        b_ldp = QPushButton('📊 Load to Preview'); b_ldp.clicked.connect(self._load_session_to_preview)
        b_dls = QPushButton('🗑 Delete Session'); b_dls.clicked.connect(self._delete_session)
        sb.addWidget(b_ldp); sb.addWidget(b_dls); sb.addStretch()
        g1.addLayout(sb)
        hlv.addWidget(gh1, 1)
        gh2 = QGroupBox('Saved Reports (in database)'); g2 = QVBoxLayout(gh2)
        self.snap_tbl = _mk_table(['ID','Report Label','Created','Period'], [46,220,160,90], stretch_col=1)
        self.snap_tbl.itemDoubleClicked.connect(lambda _: self._load_snapshot())
        g2.addWidget(self.snap_tbl)
        nb2 = QHBoxLayout()
        b_lr = QPushButton('📂 Load Report'); b_lr.clicked.connect(self._load_snapshot)
        b_ex = QPushButton('💾 Export to File'); b_ex.clicked.connect(self._export_snapshot)
        b_dl = QPushButton('🗑 Delete'); b_dl.clicked.connect(self._delete_snapshot)
        nb2.addWidget(b_lr); nb2.addWidget(b_ex); nb2.addWidget(b_dl); nb2.addStretch()
        g2.addLayout(nb2)
        hlv.addWidget(gh2, 1)
        self.tabs.addTab(tab_hist, '📋 Pull History')

        # Tab 3: Daily view (straight from DB)
        tab_daily = QWidget(); dl = QVBoxLayout(tab_daily)
        drow = QHBoxLayout()
        drow.addWidget(QLabel('Bulan:'))
        self.daily_bulan = QComboBox(); self.daily_bulan.addItems([str(i) for i in range(1,13)])
        self.daily_bulan.setCurrentText(str(now.month)); drow.addWidget(self.daily_bulan)
        drow.addWidget(QLabel('Tahun:'))
        self.daily_tahun = QComboBox()
        self.daily_tahun.addItems([str(y) for y in range(now.year-3, now.year+2)])
        self.daily_tahun.setCurrentText(str(now.year)); drow.addWidget(self.daily_tahun)
        b_dload = QPushButton('🔍 Load'); b_dload.setProperty('accent', True)
        b_dload.clicked.connect(self._refresh_daily); drow.addWidget(b_dload)
        drow.addStretch()
        dl.addLayout(drow)
        self.daily_tbl = _mk_table(
            ['Nama', 'Tanggal', 'Masuk', 'Keluar', 'Telat', 'Durasi', 'Status'],
            [130, 90, 60, 60, 55, 70, 100], stretch_col=0)
        dl.addWidget(self.daily_tbl, 1)
        self.tabs.addTab(tab_daily, '📅 Daily')

        # Tab 4: Weekly grid (Sun–Sat rows × ~4 weeks) — payroll-oriented
        tab_week = QWidget(); wkl = QVBoxLayout(tab_week)
        wrow = QHBoxLayout()
        wrow.addWidget(QLabel('Bulan:'))
        self.week_bulan = QComboBox(); self.week_bulan.addItems([str(i) for i in range(1, 13)])
        self.week_bulan.setCurrentText(str(now.month)); wrow.addWidget(self.week_bulan)
        wrow.addWidget(QLabel('Tahun:'))
        self.week_tahun = QComboBox()
        self.week_tahun.addItems([str(y) for y in range(now.year - 3, now.year + 2)])
        self.week_tahun.setCurrentText(str(now.year)); wrow.addWidget(self.week_tahun)
        wrow.addWidget(QLabel('Karyawan:'))
        self.week_emp = QComboBox(); wrow.addWidget(self.week_emp, 1)
        b_wload = QPushButton('🔍 Load'); b_wload.setProperty('accent', True)
        b_wload.clicked.connect(self._refresh_weekly); wrow.addWidget(b_wload)
        wkl.addLayout(wrow)
        tip_w = QLabel(
            'Skema payroll: baris = minggu (Minggu→Sabtu), ~4 baris / bulan. '
            'Cut-off biasanya hari Sabtu.'
        )
        tip_w.setStyleSheet('color:#888; font-size:8pt;'); tip_w.setWordWrap(True)
        wkl.addWidget(tip_w)
        self.week_tbl = QTableWidget()
        self.week_tbl.setColumnCount(8)
        self.week_tbl.setHorizontalHeaderLabels(
            ['Minggu'] + WEEKDAY_SUN_FIRST)
        self.week_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.week_tbl.verticalHeader().setVisible(False)
        self.week_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.week_tbl.setSelectionMode(QAbstractItemView.NoSelection)
        wkl.addWidget(self.week_tbl, 1)
        self.week_legend = QLabel(
            'H = Hadir (jam) · A = ABSEN · L = LIBUR · I/C/D/S = izin/cuti/dinas/sakit · NC = no checkout'
        )
        self.week_legend.setStyleSheet('color:#666; font-size:8pt;')
        wkl.addWidget(self.week_legend)
        self.tabs.addTab(tab_week, '🗓 Minggu–Sabtu')
        self._fill_week_emp_combo()

        # Status bar
        self.statusBar().showMessage('Ready.')
        self._btns = [self.btn_pull, self.btn_report, self.btn_all]

    def _mktile(self, parent_layout, caption, bg, fg):
        f = QFrame()
        f.setStyleSheet(f'background:{bg}; border-radius:10px;')
        v = QVBoxLayout(f); v.setContentsMargins(16,10,16,10)
        num = QLabel('0'); num.setAlignment(Qt.AlignCenter)
        num.setStyleSheet(f'color:{fg}; font-size:22pt; font-weight:700; background:transparent;')
        capl = QLabel(caption); capl.setAlignment(Qt.AlignCenter)
        capl.setStyleSheet(f'color:{fg}; background:transparent;')
        v.addWidget(num); v.addWidget(capl)
        parent_layout.addWidget(f)
        return num

    # ── Theme ─────────────────────────────────────────────────────────────────
    def _toggle_theme(self):
        new = 'dark' if self.cfg.get('theme','light') == 'light' else 'light'
        self.cfg['theme'] = new; save_config(self.cfg)
        self._apply_theme(new)

    def _apply_theme(self, theme):
        QApplication.instance().setStyleSheet(build_qss(theme))
        self.theme_btn.setText('☀' if theme == 'dark' else '🌙')
        self.data_lbl.setStyleSheet(f"color:{THEMES[theme]['info']}; font-size:8pt;")

    # ── Dashboards ────────────────────────────────────────────────────────────
    def _refresh_today(self):
        today = date.today()
        rows = [r for r in db_query_attendance(today.year, today.month)
                if r['timestamp'].date() == today]
        leave_map = db_leave_map(today.year, today.month)
        drows = compute_daily_rows(rows, self.cfg, leave_map)
        data, colors = [], []
        for d in drows:
            if d['leave']:
                status = d['status']
                colors.append('#DBEAFE')
            elif d['no_checkout']:
                status = 'Tanpa keluar'
                colors.append('#FEE2E2')
            elif d['telat']:
                status = f"Telat {d['telat']}m"
                colors.append('#FED7AA')
            else:
                status = 'Hadir'
                colors.append('#D1FAE5')
            if d['recovered']:
                status += ' ★'
            data.append((d['nama'], d['masuk'], status))
        _fill_table(self.today_tbl, data, colors)
        telat = sum(1 for d in drows if d['telat'] and not d['leave'])
        hadir = sum(1 for d in drows if d['status'] in ('HADIR', 'TELAT', 'NO_CHECKOUT'))
        absen = max(len(self.cfg.get('user_map', {})) - hadir, 0)
        self.tile_hadir.setText(str(hadir))
        self.tile_telat.setText(str(telat))
        self.tile_absen.setText(str(absen))

    def _fill_week_emp_combo(self):
        if not hasattr(self, 'week_emp'):
            return
        cur = self.week_emp.currentData()
        self.week_emp.blockSignals(True)
        self.week_emp.clear()
        self.week_emp.addItem('— Semua (ringkas) —', 0)
        for k, v in sorted(self.cfg.get('user_map', {}).items(), key=lambda x: int(x[0])):
            self.week_emp.addItem(f'{k} · {v}', int(k))
        if cur is not None:
            idx = self.week_emp.findData(cur)
            if idx >= 0:
                self.week_emp.setCurrentIndex(idx)
        self.week_emp.blockSignals(False)

    def _cell_bg_for_status(self, status):
        s = (status or '').upper()
        if s in ('HADIR', 'TELAT'):
            return QColor('#D1FAE5')
        if s == 'NO_CHECKOUT':
            return QColor('#FEE2E2')
        if s == 'ABSEN':
            return QColor('#FECACA')
        if s == 'LIBUR':
            return QColor('#E0E7FF')
        if s in ('IZIN', 'CUTI', 'DINAS', 'SAKIT'):
            return QColor('#DBEAFE')
        return None

    def _refresh_weekly(self):
        y = int(self.week_tahun.currentText())
        m = int(self.week_bulan.currentText())
        uid = self.week_emp.currentData()
        rows = db_query_attendance(y, m)
        leave_map = db_leave_map(y, m)
        holiday_set = db_holiday_set(y, m)
        um = {int(k): v for k, v in self.cfg.get('user_map', {}).items()}

        if uid and int(uid) != 0:
            uids = [int(uid)]
        else:
            uids = sorted(um.keys())

        # Build table: for multi-employee, stack grids with name header rows
        all_rows_data = []  # list of (is_header, label, cells)
        for u in uids:
            day_map = build_employee_day_status_map(
                u, y, m, rows, self.cfg, leave_map, holiday_set)
            weeks = build_month_week_grid(y, m, day_map)
            nama = um.get(u, f'UID:{u}')
            if len(uids) > 1:
                all_rows_data.append(('header', f'{u} · {nama}', None))
            for wk in weeks:
                all_rows_data.append(('week', wk['label'], wk['days']))

        self.week_tbl.clearContents()
        self.week_tbl.clearSpans()
        self.week_tbl.setRowCount(len(all_rows_data))
        for r, (kind, label, days) in enumerate(all_rows_data):
            if kind == 'header':
                self.week_tbl.setSpan(r, 0, 1, 8)
                it = QTableWidgetItem(label)
                f = it.font(); f.setBold(True); it.setFont(f)
                it.setBackground(QColor('#EEF2FF'))
                self.week_tbl.setItem(r, 0, it)
                self.week_tbl.setRowHeight(r, 22)
                continue
            lab = QTableWidgetItem(label)
            lab.setBackground(QColor('#F3F4F6'))
            self.week_tbl.setItem(r, 0, lab)
            for c, cell in enumerate(days):
                if not cell.get('in_month'):
                    it = QTableWidgetItem('')
                    it.setBackground(QColor('#F9FAFB'))
                else:
                    st = cell.get('status') or ''
                    txt = cell.get('text') or ''
                    # compact single-line for multi view
                    if '\n' in txt:
                        txt = txt.replace('\n', ' ')
                    it = QTableWidgetItem(txt if txt else st)
                    it.setTextAlignment(Qt.AlignCenter)
                    it.setToolTip(
                        f"{cell['date'].strftime('%Y-%m-%d')} · {st}\n"
                        f"Masuk {cell.get('masuk','-')}  Keluar {cell.get('keluar','-')}"
                    )
                    bg = self._cell_bg_for_status(st)
                    if bg:
                        it.setBackground(bg)
                    if not cell.get('in_month'):
                        pass
                self.week_tbl.setItem(r, c + 1, it)
            self.week_tbl.setRowHeight(r, 36 if len(uids) == 1 else 28)
        self._log(f'Weekly grid: {m}/{y} · {len(uids)} karyawan · {len(all_rows_data)} baris')

    def _refresh_daily(self):
        y = int(self.daily_tahun.currentText()); m = int(self.daily_bulan.currentText())
        leave_map = db_leave_map(y, m)
        drows = compute_daily_rows(db_query_attendance(y, m), self.cfg, leave_map)
        data, colors = [], []
        for d in drows:
            dur = f"{d['durasi']//60}j {d['durasi']%60}m" if d['durasi'] else '-'
            st = d['status']
            if d['recovered']:
                st += ' ★'
            data.append((d['nama'], d['tanggal'].strftime('%d-%m-%Y'), d['masuk'],
                         d['keluar'], d['telat'] or '-', dur, st))
            if d['leave']:
                colors.append('#DBEAFE')
            elif d['no_checkout']:
                colors.append('#FEE2E2')
            elif d['recovered']:
                colors.append('#E0E7FF')
            elif d['telat']:
                colors.append('#FED7AA')
            elif d['weekend']:
                colors.append('#FEF9C3')
            else:
                colors.append(None)
        _fill_table(self.daily_tbl, data, colors)
        self._log(f'Daily view: {len(drows)} baris untuk {m}/{y}')

    def _open_leave(self):
        LeaveDialog(self, self.cfg).exec()
        self._refresh_today()
        if self.tabs.currentIndex() == 3:
            self._refresh_daily()

    def _open_holiday(self):
        HolidayDialog(self).exec()

    def _export_payroll(self):
        yr = int(self.tahun_cb.currentText())
        bln = self.bulan_cb.currentText()
        _month_list = ['All', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        mo = _month_list.index(bln) if bln != 'All' else None
        if self.src_cache.isChecked():
            if not self._cache:
                QMessageBox.warning(self, 'Payroll', 'No cache. Pull data first.'); return
            rows = list(self._cache)
        else:
            rows = db_query_attendance(yr, mo)
        if mo:
            rows = [r for r in rows
                    if r['timestamp'].month == mo and r['timestamp'].year == yr]
        if not rows and not db_list_leaves(yr, mo):
            QMessageBox.warning(self, 'Payroll', 'No data for selected period.'); return
        try:
            data = generate_payroll_bytes(rows, self.cfg, yr, mo)
        except Exception as e:
            QMessageBox.critical(self, 'Payroll', str(e)); return
        period = f'{bln}_{yr}' if bln != 'All' else f'All_{yr}'
        path, _ = QFileDialog.getSaveFileName(
            self, 'Export Payroll CSV',
            f'Payroll_CVRAJ_{period}_{datetime.now().strftime("%Y%m%d")}.csv',
            'CSV (*.csv)')
        if path:
            with open(path, 'wb') as f:
                f.write(data)
            self._log(f'✓ Payroll CSV → {path}')
            _open_path(os.path.dirname(path))

    # ── UI helpers ────────────────────────────────────────────────────────────
    def _log(self, msg):
        def do():
            ts = datetime.now().strftime('%H:%M:%S')
            self.log_box.appendPlainText(f'[{ts}] {msg}')
            self.statusBar().showMessage(msg)
        self.ui(do)   # safe from any thread

    def _get_conn(self):
        from zk import ZK
        try:
            pw = int(self.cfg.get('comm_password', 0) or 0)
        except (TypeError, ValueError):
            pw = 0
        return ZK(self.ip_edit.text().strip(), port=int(self.port_edit.text()),
                  timeout=15, password=pw, force_udp=False, ommit_ping=False).connect()

    def _set_buttons(self, enabled):
        for b in self._btns: b.setEnabled(enabled)

    def _run(self, fn):
        self._set_buttons(False)
        threading.Thread(target=self._worker, args=(fn, False), daemon=True).start()

    def _run_quiet(self, fn):
        """Background job without modal error dialogs (tray + log only)."""
        self._set_buttons(False)
        threading.Thread(target=self._worker, args=(fn, True), daemon=True).start()

    def _worker(self, fn, quiet=False):
        try:
            fn()
        except Exception as e:
            self._log(f'[ERROR] {e}')
            if quiet:
                self.ui(lambda e=e: self._tray_msg(
                    'ZKTeco Error', str(e), QSystemTrayIcon.Warning))
            else:
                self.ui(lambda e=e: QMessageBox.critical(self, 'Error', str(e)))
        finally:
            self.ui(lambda: self._set_buttons(True))
            self.ui(self._update_status)

    def _update_status(self):
        n = db_count()
        self.data_lbl.setText(f'DB: {n:,} records  |  Cache: {len(self._cache):,}')

    def _open_settings(self):
        def on_save(new_cfg):
            self.cfg = new_cfg
            self.ip_edit.setText(new_cfg['ip'])
            self.port_edit.setText(str(new_cfg['port']))
            enable = bool(new_cfg.get('autostart') or new_cfg.get('silent_mode'))
            self._apply_autostart(enable)
            # reschedule interval
            if getattr(self, '_pull_timer', None):
                self._pull_timer.stop()
                self._pull_timer = None
            try:
                mins = int(new_cfg.get('auto_pull_interval_min', 0) or 0)
            except (TypeError, ValueError):
                mins = 0
            if mins > 0:
                self._pull_timer = QTimer(self)
                self._pull_timer.timeout.connect(self._startup_auto_pull)
                self._pull_timer.start(mins * 60 * 1000)
                self._log(f'Auto-pull interval: setiap {mins} menit')
            self._log('✓ Settings saved')
        SettingsDialog(self, self.cfg, on_save).exec()

    def _apply_autostart(self, enable):
        """Register/unregister in HKCU Run — launches with --silent at login."""
        if sys.platform != 'win32':
            return
        import winreg
        try:
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r'Software\Microsoft\Windows\CurrentVersion\Run',
                0, winreg.KEY_SET_VALUE,
            )
            if enable:
                if getattr(sys, 'frozen', False):
                    cmd = f'"{sys.executable}" --silent'
                else:
                    # dev: pythonw preferred if available, else python
                    py = sys.executable
                    script = os.path.abspath(__file__)
                    cmd = f'"{py}" "{script}" --silent'
                winreg.SetValueEx(key, 'ZKTeco_Utility', 0, winreg.REG_SZ, cmd)
                self._log('✓ Autostart ON — login Windows → silent tray + auto pull')
                self._log(f'  Run key: {cmd}')
            else:
                try:
                    winreg.DeleteValue(key, 'ZKTeco_Utility')
                except FileNotFoundError:
                    pass
                self._log('✓ Autostart OFF')
            key.Close()
        except Exception as e:
            self._log(f'[ERROR] Autostart: {e}')

    def _toast(self, msg):
        """Notifikasi punch. Saat window disembunyikan ke tray → balloon tray saja
        (hindari 'jendela hantu' floating yang sering dianggap glitch)."""
        if not self.isVisible() or self.isMinimized():
            self._tray_msg('Absensi', msg)
            return
        t = QLabel(msg, self)  # child of main window → tidak jadi top-level window
        t.setWindowFlags(Qt.Tool | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        t.setAttribute(Qt.WA_ShowWithoutActivating, True)
        t.setStyleSheet(f'background:{ACCENT}; color:white; font-size:10pt; font-weight:600; '
                        'padding:12px 18px; border-radius:10px;')
        t.adjustSize()
        scr = QApplication.primaryScreen().availableGeometry()
        # posisi global, tapi jangan curi focus dari app lain
        t.move(scr.right() - t.width() - 16, scr.bottom() - t.height() - 16)
        eff = QGraphicsOpacityEffect(t); t.setGraphicsEffect(eff)
        anim = QPropertyAnimation(eff, b'opacity', t)
        anim.setDuration(250); anim.setStartValue(0); anim.setEndValue(1)
        anim.setEasingCurve(QEasingCurve.OutCubic); anim.start()
        t.show()
        self._toasts.append(t)   # keep a ref so it isn't GC'd
        QTimer.singleShot(4000, lambda: (t.close(), self._toasts.remove(t) if t in self._toasts else None))

    def _on_lang_change(self, lang):
        self.cfg['lang'] = lang; save_config(self.cfg)
        # lang only affects generated reports — applies on next Preview, no restart needed
        self._log(f'✓ Report language: {"English" if lang == "en" else "Indonesia"}')

    def _confirm_clear(self):
        if QMessageBox.question(self, 'Confirm',
                'Delete ALL attendance log from device memory?\n\n'
                'Data already in local database will NOT be deleted.') == QMessageBox.Yes:
            self._run(self._do_clear)

    # ── Preview / Viewer ──────────────────────────────────────────────────────
    def _load_excel_to_viewer(self, data_bytes, label=''):
        try:
            from openpyxl import load_workbook
            import io
            wb = load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            self._current_snap_bytes = data_bytes
            self.sheet_cb.clear(); self.sheet_cb.addItems(sheet_names)
            self.snap_lbl.setText(label or 'Report loaded')
            self._render_sheet(data_bytes, 0)
        except Exception as e:
            self._log(f'[ERROR] Cannot load preview: {e}')

    def _render_sheet(self, data_bytes, sheet_idx):
        headers, rows = parse_excel_for_preview(data_bytes, sheet_idx)
        def show():
            self.report_tbl.setColumnCount(len(headers))
            self.report_tbl.setHorizontalHeaderLabels([h.replace('\n',' ') for h in headers])
            _fill_table(self.report_tbl, rows)
            self.report_tbl.resizeColumnsToContents()
            for i in range(len(headers)):   # cap width so day-matrix sheets stay compact
                if self.report_tbl.columnWidth(i) > 200: self.report_tbl.setColumnWidth(i, 200)
        self.ui(show)

    def _on_sheet_change(self, idx):
        if not self._current_snap_bytes: return
        threading.Thread(target=lambda: self._render_sheet(self._current_snap_bytes, idx),
                         daemon=True).start()

    def _save_excel(self):
        if not self._current_snap_bytes:
            QMessageBox.warning(self, 'No Report', 'Generate a report first.')
            return
        path, _ = QFileDialog.getSaveFileName(self, 'Save Excel',
            f'Absensi_CVRAJ_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx', 'Excel (*.xlsx)')
        if path:
            with open(path, 'wb') as f: f.write(self._current_snap_bytes)
            self._log(f'✓ Saved to {path}')
            _open_path(os.path.dirname(path))

    # ── History ───────────────────────────────────────────────────────────────
    def _refresh_history(self):
        sess = db_get_pull_sessions()
        _fill_table(self.sess_tbl,
                    [(sid, pulled_at, f'{rec:,}', f'+{new_:,}', ip)
                     for sid, pulled_at, rec, new_, ip in sess])
        snaps = db_get_excel_snapshots()
        _fill_table(self.snap_tbl,
                    [(sid, label, created_at, f'{yr}/{mo:02d}' if mo else str(yr))
                     for sid, label, created_at, yr, mo, _, _ in snaps])

    def _sel_id(self, tbl):
        r = tbl.currentRow()
        return int(tbl.item(r, 0).text()) if r >= 0 else None

    def _load_session_to_preview(self):
        sid = self._sel_id(self.sess_tbl)
        if sid is None:
            QMessageBox.warning(self, 'Select', 'Select a pull session first.'); return
        conn = sqlite3.connect(DB_FILE); c = conn.cursor()
        c.execute('SELECT pulled_at FROM pull_sessions WHERE id=?', (sid,))
        row = c.fetchone(); conn.close()
        if not row: return
        pulled_at = row[0]
        conn = sqlite3.connect(DB_FILE); c = conn.cursor()
        c.execute('SELECT uid,nama,timestamp,punch FROM attendance WHERE pulled_at=? ORDER BY timestamp', (pulled_at,))
        att_rows = [{'uid': r[0], 'nama': r[1],
                     'timestamp': datetime.strptime(r[2], '%Y-%m-%d %H:%M:%S'), 'punch': r[3]}
                    for r in c.fetchall()]
        conn.close()
        if not att_rows:
            QMessageBox.information(self, 'Empty', 'No attendance records found for this session.'); return
        self._run(lambda: self._generate_and_show(att_rows, f'Session #{sid} ({pulled_at[:10]})', sid))

    def _generate_and_show(self, rows, label, session_id=None):
        self._log(f'Generating report for {label} ...')
        yr = int(self.tahun_cb.currentText())
        bln = self.bulan_cb.currentText()
        mo = (['All','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'].index(bln)
              if bln != 'All' else None)
        if mo: rows = [r for r in rows if r['timestamp'].month == mo and r['timestamp'].year == yr]
        data = generate_excel_bytes(rows, self.cfg)
        period = f'{bln} {yr}' if bln != 'All' else f'All {yr}'
        snap_id = db_save_excel_snapshot(session_id or 0, f'{label} — {period}', yr, mo, data)
        self._current_snap_id = snap_id
        self.ui(lambda: self._load_excel_to_viewer(data, f'{label} — {period}'))
        self.ui(self._refresh_history)
        self._log(f'✓ Report ready — {len(rows)} records | Saved to DB (snapshot #{snap_id})')

    def _load_snapshot(self):
        snap_id = self._sel_id(self.snap_tbl)
        if snap_id is None:
            QMessageBox.warning(self, 'Select', 'Select a report first.'); return
        row = db_load_excel_snapshot(snap_id)
        if not row: return
        data, label = row
        self._current_snap_id = snap_id
        self._load_excel_to_viewer(data, label)
        self.tabs.setCurrentIndex(1)
        self._log(f'✓ Loaded snapshot #{snap_id}: {label}')

    def _export_snapshot(self):
        snap_id = self._sel_id(self.snap_tbl)
        if snap_id is None:
            if self._current_snap_bytes: self._save_excel(); return
            QMessageBox.warning(self, 'Select', 'Select a report to export.'); return
        row = db_load_excel_snapshot(snap_id)
        if not row: return
        data, label = row
        safe = label.replace(' ', '_').replace('/', '_').replace(':', '')[:40]
        path, _ = QFileDialog.getSaveFileName(self, 'Export Excel', f'{safe}.xlsx', 'Excel (*.xlsx)')
        if path:
            with open(path, 'wb') as f: f.write(data)
            self._log(f'✓ Exported to {path}')
            _open_path(os.path.dirname(path))

    def _delete_session(self):
        sid = self._sel_id(self.sess_tbl)
        if sid is None: return
        if QMessageBox.question(self, 'Delete',
                f'Delete pull session #{sid}?\nAll saved reports for this session will also be deleted.') == QMessageBox.Yes:
            db_delete_pull_session(sid)
            self._refresh_history()
            self._log(f'✓ Session #{sid} deleted')

    def _delete_snapshot(self):
        snap_id = self._sel_id(self.snap_tbl)
        if snap_id is None: return
        if QMessageBox.question(self, 'Delete', f'Delete report snapshot #{snap_id}?') == QMessageBox.Yes:
            db_delete_excel_snapshot(snap_id)
            self._refresh_history()
            self._log(f'Snapshot #{snap_id} deleted')

    # ── Device actions ────────────────────────────────────────────────────────
    def _do_test(self):
        ip = self.ip_edit.text().strip()
        self._log(f'Testing connection to {ip}:{self.port_edit.text()} ...')
        conn = self._get_conn()
        try:
            fw = conn.get_firmware_version()
            self._log(f'✓ Connected! Firmware: {fw}')
            self.ui(lambda: self.conn_lbl.setStyleSheet('color:#16a34a; font-size:8pt;'))
            self.ui(lambda: self.conn_lbl.setText(f'● {ip}'))
            self._check_clock(conn)
            self._check_capacity(conn)
        finally: conn.disconnect()

    def _do_info(self):
        self._log('Fetching device info ...')
        conn = self._get_conn()
        try:
            conn.read_sizes()
            info = {'Serial Number': conn.get_serialnumber(), 'Firmware': conn.get_firmware_version(),
                    'Device Time': str(conn.get_time()),
                    'Users': f'{conn.users} / {conn.users_cap}',
                    'Logs': f'{conn.records:,} / {conn.rec_cap:,}',
                    'Local DB': f'{db_count():,} records'}
            self._log('✓ Device info received')
            self.ui(lambda: DeviceInfoDialog(self, info).exec())
        finally: conn.disconnect()

    def _sync_user_map_from_device(self, ulist):
        """Rebuild user_map from live device list — drop resign/ghost UIDs."""
        old = dict(self.cfg.get('user_map') or {})
        new_map = {}
        for u in ulist:
            key = str(int(u['uid']))
            # keep config name if device name empty
            name = (u.get('nama') or '').strip() or old.get(key) or f"UID:{key}"
            new_map[key] = name
        pruned = sorted(set(old) - set(new_map), key=lambda x: int(x) if x.isdigit() else x)
        self.cfg['user_map'] = new_map
        save_config(self.cfg)
        if pruned:
            names = ', '.join(f"{k}:{old[k]}" for k in pruned)
            self._log(f'  user_map pruned (not on device): {names}')
        self.ui(self._fill_week_emp_combo)
        return new_map

    def _do_users(self):
        self._log('Fetching users from device ...')
        conn = self._get_conn()
        try:
            users = conn.get_users()
            ulist = [{'uid': int(u.user_id), 'nama': u.name, 'card_id': getattr(u, 'card', '') or ''} for u in users]
            db_upsert_users(ulist)
            self._sync_user_map_from_device(ulist)
            self._log(f'✓ {len(ulist)} users on device (user_map={len(self.cfg["user_map"])})')
            self.ui(lambda: UserManagerDialog(self, ulist, app=self).exec())
        finally: conn.disconnect()

    def _device_user_op(self, op, dlg, done_msg):
        """Run op(conn) on the device, then re-fetch users into the dialog."""
        conn = self._get_conn()
        try:
            op(conn)
            users = conn.get_users()
            ulist = [{'uid': int(u.user_id), 'nama': u.name, 'card_id': getattr(u, 'card', '') or ''} for u in users]
            db_upsert_users(ulist)
            self._sync_user_map_from_device(ulist)
            self._log(done_msg + f'  | device users: {len(ulist)}')
            self.ui(lambda: dlg._fill(ulist))
        finally: conn.disconnect()

    def device_user_save(self, uid, name, dlg):
        def op(conn):
            ex = next((u for u in conn.get_users() if str(u.user_id) == uid), None)
            if ex:  # rename, keep everything else (privilege, card, password)
                conn.set_user(uid=ex.uid, name=name, privilege=ex.privilege,
                              password=ex.password or '', group_id=ex.group_id or '',
                              user_id=uid, card=ex.card or 0)
            else:
                conn.set_user(name=name, user_id=uid)
        self._run(lambda: self._device_user_op(op, dlg, f'✓ User {uid} = {name} saved to device'))

    def device_user_delete(self, uid, dlg):
        self._run(lambda: self._device_user_op(
            lambda conn: conn.delete_user(user_id=uid), dlg, f'✓ User {uid} deleted from device'))

    def _confirm_restart(self):
        if QMessageBox.question(self, 'Restart',
                'Restart the device now?\nIt will be offline for ~1 minute.') == QMessageBox.Yes:
            self._run(self._do_restart)

    # ── Live monitor ──────────────────────────────────────────────────────────
    # Holds its own connection open; not routed through _run so the other
    # buttons stay usable while monitoring.
    def _toggle_live(self):
        if getattr(self, '_live_want', False):
            self._live_want = False
            c = getattr(self, '_live_conn', None)
            if c: c.end_live_capture = True   # loop in _live_loop exits
            return
        self._live_want = True
        threading.Thread(target=self._live_loop, daemon=True).start()

    def _live_loop(self):
        self.ui(lambda: self.btn_live.setText('⏹ Stop Live'))
        self._log('📡 Live monitor ON — punches appear here as they happen')
        while self._live_want:   # reconnect loop — survives device/network drops
            conn = None
            try:
                conn = self._get_conn()
                self._live_conn = conn
                self._check_clock(conn, quiet=True)   # power-loss guard on every (re)connect
                clock_at = time.time()
                um = {int(k): v for k, v in self.cfg.get('user_map', {}).items()}
                for att in conn.live_capture():
                    if att is None:   # idle timeout tick
                        if not self._live_want: break
                        if time.time() - clock_at > 600:   # ponytail: re-sync via reconnect cycle
                            conn.end_live_capture = True; break
                        continue
                    name = um.get(int(att.user_id), f'UID:{att.user_id}')
                    hhmm = att.timestamp.strftime('%H:%M') if att.timestamp else '?'
                    self._log(f'👆 {name} — {att.timestamp}')
                    self.ui(lambda n=name, h=hhmm: self._toast(f'👆 {n} absen — {h}'))
                    # save immediately; UNIQUE(uid,timestamp) dedups later pulls
                    ins, skip = db_insert_attendance([{
                        'uid': int(att.user_id), 'nama': name,
                        'timestamp': att.timestamp, 'punch': att.punch,
                    }])
                    if skip:
                        self._log(f'  (skip dup punch {name} @ {att.timestamp})')
                    self.ui(self._update_status)
                    self.ui(self._refresh_today)
            except Exception as e:
                if self._live_want: self._log(f'⚠ Live monitor: {e} — reconnecting in 30s')
            finally:
                self._live_conn = None
                try:
                    if conn: conn.disconnect()
                except Exception: pass
            for _ in range(30):   # wait, but stay responsive to Stop
                if not self._live_want: break
                time.sleep(1)
        self.ui(lambda: self.btn_live.setText('📡 Live Monitor'))
        self._log('📡 Live monitor OFF')

    def _check_capacity(self, conn, quiet=False):
        """Warn when the device log memory is nearly full."""
        try:
            conn.read_sizes()
            if conn.rec_cap and conn.records / conn.rec_cap >= 0.8:
                pct = round(conn.records / conn.rec_cap * 100)
                self._log(f'⚠ Device log {pct}% full ({conn.records:,}/{conn.rec_cap:,})')
                msg = (
                    f'Memori log mesin {pct}% penuh ({conn.records:,}/{conn.rec_cap:,}). '
                    f'Pull + Clear Device Log.'
                )
                if quiet:
                    self.ui(lambda: self._tray_msg('Log Hampir Penuh', msg, QSystemTrayIcon.Warning))
                else:
                    self.ui(lambda: QMessageBox.warning(self, 'Log Hampir Penuh',
                        f'Memori log mesin {pct}% penuh ({conn.records:,} dari {conn.rec_cap:,}).\n\n'
                        f'Pull Data dulu, lalu jalankan "Clear Device Log" — kalau penuh, '
                        f'absensi baru tidak akan tersimpan.'))
        except Exception as e:
            self._log(f'  (capacity check skipped: {e})')

    def _do_restart(self):
        self._log('Restarting device ...')
        conn = self._get_conn()
        conn.restart()   # device drops the link; no disconnect() after this
        self._log('✓ Restart command sent — device back in ~1 minute')

    def _do_pull(self, quiet=False):
        self._log('Pulling attendance data from device ...')
        conn = self._get_conn()
        try:
            self._check_clock(conn, quiet=quiet)
            self._check_capacity(conn, quiet=quiet)
            atts = conn.get_attendance()
            if not atts:
                self._log('⚠ No data on device.')
                if quiet:
                    self.ui(lambda: self._tray_msg('ZKTeco Pull', 'Tidak ada data di mesin.'))
                return
            um = {int(k): v for k, v in self.cfg.get('user_map', {}).items()}
            recs = [{'uid': int(a.user_id), 'nama': um.get(int(a.user_id), f'UID:{a.user_id}'),
                     'timestamp': a.timestamp, 'punch': a.punch} for a in atts if a.timestamp]
            anomaly = [r for r in recs if is_anomaly_ts(r['timestamp'])]
            normal  = [r for r in recs if not is_anomaly_ts(r['timestamp'])]
            if anomaly and self.cfg.get('anomaly_recover', True):
                cfg_anchor = str(self.cfg.get('anomaly_anchor', '') or '').strip()
                anchor = None
                if cfg_anchor:
                    try:
                        anchor = datetime.strptime(cfg_anchor, '%Y-%m-%d').date()
                    except Exception:
                        anchor = None
                gaps = find_gaps(normal, min_len=2)
                if anchor is None:
                    anchor = find_gap_start(normal)
                if anchor is None:
                    # multi-gap or no calendar — require manual anchor
                    gap_txt = '\n'.join(
                        f'  • {s.strftime("%d %b %Y")} ({n} hari)' for s, n in gaps[:5]
                    ) or '  (tidak ada gap terdeteksi di data normal)'
                    self._log(f'⚠ {len(anomaly)} anomaly records — recovery NEEDS manual anchor date.')
                    self._log(f'  Significant gaps:\n{gap_txt}')
                    if quiet:
                        self.ui(lambda: self._tray_msg(
                            'Butuh Recovery Anchor',
                            f'{len(anomaly)} record anomaly — set anchor di Settings.',
                            QSystemTrayIcon.Warning))
                    else:
                        self.ui(lambda: QMessageBox.warning(
                            self, 'Set Recovery Anchor',
                            f'{len(anomaly)} record tahun {ANOMALY_YEAR} ditemukan, tetapi '
                            f'auto-remap TIDAK dijalankan.\n\n'
                            f'Alasan: beberapa gap multi-hari atau tidak ada data normal — '
                            f'risiko menempel di tanggal libur/cuti.\n\n'
                            f'Buka Settings → isi "Recovery anchor date" (hari pertama outage), '
                            f'simpan, lalu Pull lagi.\n\n'
                            f'Gap terdeteksi:\n{gap_txt}'))
                    self._cache = normal  # store only safe rows
                else:
                    remapped = remap_anomalies(
                        anomaly, anchor,
                        self.cfg.get('jam_masuk', '08:00'),
                        self.cfg.get('jam_keluar', '16:00'))
                    n_days = len(set(r['timestamp'].date() for r in anomaly))
                    last = anchor + timedelta(days=max(0, n_days - 1))
                    self._log(f'⚠ {len(anomaly)} ANOMALY records (clock reset year {ANOMALY_YEAR}).')
                    self._log(f'  → recovered onto {anchor.strftime("%d %b")} .. '
                              f'{last.strftime("%d %b %Y")} ({len(remapped)} punches) [★ recovered]')
                    if len(gaps) > 1:
                        self._log(f'  (note: {len(gaps)} gaps found; using anchor {anchor})')
                    self._backup_anomaly_csv(remapped)
                    if quiet:
                        self.ui(lambda: self._tray_msg(
                            'Anomaly Recovered',
                            f'{len(anomaly)} record dipulihkan ★ '
                            f'{anchor.strftime("%d %b")}–{last.strftime("%d %b %Y")}'))
                    else:
                        self.ui(lambda: QMessageBox.warning(
                            self, 'Anomaly Recovered',
                            f'{len(anomaly)} record jam ter-reset DIPULIHKAN (ditandai ★).\n\n'
                            f'{anchor.strftime("%d %b %Y")} s/d {last.strftime("%d %b %Y")}\n\n'
                            f'Backup audit CSV di folder aplikasi.'))
                    self._cache = normal + remapped
            else:
                if anomaly:
                    self._log(f'⚠ {len(anomaly)} anomaly records ignored (recovery disabled).')
                self._cache = normal
            new, skipped = db_insert_attendance(self._cache)
            sid = db_add_pull_session(len(self._cache), new, self.ip_edit.text().strip())
            self._log(f'✓ {len(atts)} records pulled  |  {new} new saved  |  {skipped} skipped (dup)')
            if skipped:
                self._log(f'  (dup = same UID+timestamp already in DB)')
            self._log(f'  Pull session #{sid} recorded')
            if self.cfg.get('auto_backup', False):
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                path = os.path.join(_BASE, f'backup_raw_{ts}.csv')
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    w = csv.writer(f)
                    w.writerow(['UserID', 'Name', 'Timestamp', 'Punch', 'Recovered'])
                    for r in self._cache:
                        w.writerow([
                            r['uid'], r['nama'],
                            r['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if r['timestamp'] else '',
                            r.get('punch', 0),
                            'Y' if r.get('recovered') else '',
                        ])
                self._log(f'  Backup CSV → {path}')
            # Auto cloud sync after pull (if enabled)
            sync_note = ''
            if self.cfg.get('cloud_sync_enabled') and self.cfg.get('cloud_sync_after_pull', True):
                try:
                    cloud_sync(self.cfg, punches=self._cache, log=self._log)
                    sync_note = ' + cloud OK'
                except Exception as e:
                    self._log(f'⚠ Cloud sync gagal: {e}')
                    sync_note = f' + cloud GAGAL: {e}'
                    if quiet:
                        self.ui(lambda e=e: self._tray_msg(
                            'Cloud Sync Gagal', str(e), QSystemTrayIcon.Warning))
            if quiet:
                self.ui(lambda n=new, s=skipped, sn=sync_note: self._tray_msg(
                    'ZKTeco Pull',
                    f'{n} baru, {s} dup{sn}',
                ))
            self.ui(self._refresh_history)
            self.ui(self._refresh_today)
        finally:
            conn.disconnect()

    def _do_cloud_sync(self):
        """Manual full DB sync → VST (employees + all punches + leaves)."""
        try:
            cloud_sync(self.cfg, punches=None, log=self._log)
            self.ui(lambda: QMessageBox.information(
                self, 'Cloud Sync', 'Sinkron ke VST berhasil.'))
        except Exception as e:
            self._log(f'⚠ Cloud sync gagal: {e}')
            self.ui(lambda: QMessageBox.warning(self, 'Cloud Sync', str(e)))

    def _check_clock(self, conn, quiet=False):
        """Auto-sync device RTC to the PC clock when it has drifted."""
        try:
            dev = conn.get_time(); pc = datetime.now()
            skew = abs((dev - pc).total_seconds())
            if skew > 120:
                mins = int(skew // 60)
                conn.set_time(datetime.now())
                self._log(f'⚠ Device clock was off by ~{mins} min (was {dev}) — auto-synced to PC time')
            elif not quiet:
                self._log(f'✓ Device clock OK ({dev})')
            return skew
        except Exception as e:
            if not quiet: self._log(f'  (clock check skipped: {e})')
            return None

    def _backup_anomaly_csv(self, remapped):
        """Save original (corrupted) vs remapped timestamps for audit."""
        try:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            path = os.path.join(_BASE, f'anomaly_recovered_{ts}.csv')
            with open(path, 'w', newline='', encoding='utf-8') as f:
                w = csv.writer(f); w.writerow(['UID','Name','OriginalTimestamp','RemappedTimestamp','Status'])
                for r in sorted(remapped, key=lambda x: x['timestamp']):
                    w.writerow([r['uid'], r['nama'],
                                r['orig_ts'].strftime('%Y-%m-%d %H:%M:%S'),
                                r['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                                'Check-In' if r['punch'] == 0 else 'Check-Out'])
            self._log(f'  Raw anomaly backup → {path}')
        except Exception as e:
            self._log(f'  (anomaly backup failed: {e})')

    def _do_clear(self):
        self._log('Clearing log from device memory ...')
        conn = self._get_conn()
        try:
            conn.clear_attendance()
            self._log('✓ Device attendance log cleared')
        finally: conn.disconnect()

    def _do_report(self):
        yr = int(self.tahun_cb.currentText())
        bln = self.bulan_cb.currentText()
        src = 'cache' if self.src_cache.isChecked() else 'database'
        _month_list = ['All','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        mo = _month_list.index(bln) if bln != 'All' else None
        if src == 'cache':
            if not self._cache: raise RuntimeError('No cache. Pull data first.')
            rows = self._cache
        else:
            rows = db_query_attendance(yr, mo)
            if not rows: raise RuntimeError(f'No data in database for {bln} {yr}.')
        label = f"{'DB' if src == 'database' else 'Cache'} {bln} {yr}"
        self._generate_and_show(rows, label)
        self.ui(lambda: self.tabs.setCurrentIndex(1))

    def _do_all(self):
        self._do_pull()   # clock auto-syncs inside pull
        self._do_report()

    # ── Updater ───────────────────────────────────────────────────────────────
    def _check_update(self):
        try:
            from updater import get_latest_release, is_newer
        except ImportError:
            QMessageBox.information(self, 'Update', 'Updater module not found.'); return
        self._log('Checking for updates on GitHub...')
        def worker():
            info = get_latest_release()
            if info is None:
                self._log('Cannot reach GitHub. Check internet.')
                return
            latest = info['version']; dl_url = info['download_url']; body = info.get('body', '') or ''
            if not is_newer(latest, APP_VERSION):
                self._log(f'Already up to date (v{APP_VERSION})')
                self.ui(lambda: QMessageBox.information(self, 'Up to Date',
                    f'You have the latest version.\nCurrent: v{APP_VERSION}'))
                return
            self.ui(lambda: self._prompt_update(latest, dl_url, body))
        threading.Thread(target=worker, daemon=True).start()

    def _prompt_update(self, version, dl_url, body):
        changelog = body[:500] if body else '(no changelog)'
        msg = (f'New version available: {version}\n'
               f'Current: v{APP_VERSION}\n\n'
               f'Changelog:\n{changelog}\n\n'
               f'Download and install now?\n'
               f'App will restart automatically.')
        if QMessageBox.question(self, 'Update Available', msg) == QMessageBox.Yes:
            self._do_download(dl_url, version)

    def _do_download(self, url, version):
        try:
            from updater import download_and_replace
        except ImportError:
            QMessageBox.critical(self, 'Error', 'Updater module not found.'); return
        self._log(f'Starting download of v{version} ...')
        dlg = QDialog(self)
        dlg.setWindowTitle(f'Updating to {version}'); dlg.setModal(True)
        dlg.setWindowFlags(dlg.windowFlags() & ~Qt.WindowCloseButtonHint)
        v = QVBoxLayout(dlg)
        t1 = QLabel('ZKTeco eFace10 Utility'); t1.setStyleSheet('font-size:11pt; font-weight:700;')
        t2 = QLabel(f'Updating to {version}'); t2.setStyleSheet('color:#555;')
        t3 = QLabel('Do not close this window.'); t3.setStyleSheet('color:#888; font-size:8pt;')
        v.addWidget(t1, alignment=Qt.AlignCenter); v.addWidget(t2, alignment=Qt.AlignCenter)
        v.addWidget(t3, alignment=Qt.AlignCenter)
        pbar = QProgressBar(); pbar.setRange(0, 100); pbar.setFixedWidth(360); v.addWidget(pbar)
        pct_lbl = QLabel('0%'); pct_lbl.setStyleSheet(f'color:{ACCENT}; font-weight:700;')
        v.addWidget(pct_lbl, alignment=Qt.AlignCenter)
        step_lbl = QLabel('Connecting to GitHub...'); step_lbl.setStyleSheet('color:#555; font-size:8pt;')
        v.addWidget(step_lbl, alignment=Qt.AlignCenter)
        dlg.show()

        def on_progress(pct):
            self.ui(lambda: (pbar.setValue(pct), pct_lbl.setText(f'{pct}%')))
        def on_status(msg):
            self.ui(lambda: step_lbl.setText(msg))
            self._log(f'  {msg}')
        def on_done():
            self.ui(lambda: (pbar.setValue(100), pct_lbl.setText('100%'),
                             step_lbl.setText('Done! Restarting in 2 seconds...')))
            self._log(f'Update {version} installed successfully')
            self.ui(lambda: QTimer.singleShot(2000, lambda: self._finish_update(dlg)))
        def on_error(msg):
            self.ui(dlg.close)
            self._log(f'[ERROR] Update failed: {msg}')
            self.ui(lambda: QMessageBox.critical(self, 'Update Failed',
                f'Failed to install update:\n\n{msg}\n\n'
                f'Download manually:\ngithub.com/xbanana29/zkteco-utility/releases'))

        download_and_replace(url,
            on_progress=on_progress,
            on_status=on_status,
            on_done=on_done,
            on_error=on_error)

    def _finish_update(self, dlg):
        try: dlg.close()
        except Exception: pass
        try:
            from updater import restart_app
            restart_app()
        except Exception:
            QApplication.quit()


def _want_silent_boot(cfg=None):
    if '--show' in sys.argv:
        return False
    if '--silent' in sys.argv or '--minimized' in sys.argv:
        return True
    if cfg is not None and cfg.get('silent_mode'):
        return True
    return False


def _handoff_to_running_instance():
    """Jika instance sudah jalan, minta buka window (kecuali boot silent). Return True = exit."""
    sock = QLocalSocket()
    sock.connectToServer(_INSTANCE_KEY)
    if not sock.waitForConnected(250):
        return False
    cmd = b'ping\n' if _want_silent_boot() else b'show\n'
    sock.write(cmd)
    sock.flush()
    sock.waitForBytesWritten(400)
    sock.disconnectFromServer()
    return True


def main():
    qapp = QApplication(sys.argv)   # theme stylesheet applied by App._apply_theme
    qapp.setQuitOnLastWindowClosed(False)   # closing to tray must not quit

    # Single instance: cegah window "muncul sendiri" karena autostart + buka manual
    # (atau dua shortcut) membuka proses kedua.
    if _handoff_to_running_instance():
        return 0

    QLocalServer.removeServer(_INSTANCE_KEY)
    server = QLocalServer()
    if not server.listen(_INSTANCE_KEY):
        # stale lock — coba sekali lagi
        QLocalServer.removeServer(_INSTANCE_KEY)
        server.listen(_INSTANCE_KEY)

    win = App()

    def _on_ipc():
        client = server.nextPendingConnection()
        if not client:
            return
        client.waitForReadyRead(300)
        data = bytes(client.readAll()).decode('utf-8', errors='ignore').strip().lower()
        client.disconnectFromServer()
        # show = buka dashboard; ping = instance kedua silent, abaikan
        if data.startswith('show') or data == '':
            win._tray_open()

    server.newConnection.connect(_on_ipc)

    silent = _want_silent_boot(win.cfg)
    if silent and win._tray:
        # stay hidden in tray; auto-pull scheduled in App.__init__
        pass
    else:
        win.show()
    sys.exit(qapp.exec())


if __name__ == '__main__':
    main()
